"""
TCE Financial Statement Engine - Comprehensive E2E Test
Tests ALL modules across all 5 sprints:
  Sprint 1: Client, Project, TB Upload, Mapping, Audit
  Sprint 2: Financial Engine (BS+PL), Notes, Cash Flow, Ratios, EPS
  Sprint 3: Excel Export, Ageing, Share Capital
  Sprint 4: Web Frontend pages
  Sprint 5: Related Party, Disclosures, Accounting Policies
"""
import io, json, sys
from fastapi.testclient import TestClient
from app.main import app

PASS = 0
FAIL = 0

def check(label, condition, detail=""):
    global PASS, FAIL
    if condition:
        PASS += 1
        print(f"  [PASS] {label}")
    else:
        FAIL += 1
        print(f"  [FAIL] {label} {detail}")


with TestClient(app) as c:
    print("=" * 70)
    print("TCE ENGINE - COMPREHENSIVE E2E TEST")
    print("=" * 70)

    # ========== SPRINT 1: FOUNDATION ==========
    print("\n--- SPRINT 1: Foundation ---")

    # CoA seeding
    r = c.get("/api/coa/")
    check("CoA codes loaded", r.status_code == 200 and len(r.json()) == 248,
          f"got {len(r.json()) if r.status_code == 200 else 'error'}")

    r = c.get("/api/coa/?fs_type=BS")
    check("BS codes count", len(r.json()) == 173)

    r = c.get("/api/coa/?fs_type=PL")
    check("PL codes count", len(r.json()) == 75)

    # Client CRUD
    r = c.post("/api/clients/", json={
        "name": "Evenset Consultancy Services OPC Pvt Ltd",
        "cin": "U74999KA2020OPC000001",
        "auditor_name": "ABC & Associates", "auditor_frn": "123456W",
        "face_value": 10, "paid_up_capital": 100000,
        "authorized_capital": 1000000, "tax_rate": 0.2522,
    })
    check("Create client", r.status_code == 200 and r.json()["id"] == 1)

    r = c.get("/api/clients/1")
    check("Get client detail", r.status_code == 200 and "Evenset" in r.json()["name"])

    # Project CRUD
    r = c.post("/api/projects/", json={
        "client_id": 1, "financial_year": "2024-25",
        "bs_date_cy": "2025-03-31", "bs_date_py": "2024-03-31", "rounding": "Rupees"
    })
    check("Create project", r.status_code == 200 and r.json()["id"] == 1)

    # TB Upload
    sample_tb = """Ledger,Tally Group,CY Debit,CY Credit,PY Debit,PY Credit
Share Capital,Capital Account,0,100000,0,100000
Securities Premium,Reserves & Surplus,0,50000,0,50000
Profit & Loss A/c,Reserves & Surplus,0,25000,0,5000
Term Loan,Secured Loans,0,500000,0,600000
Sundry Creditor - ABC,Sundry Creditors,0,45000,0,40000
Sundry Creditor - XYZ,Sundry Creditors,0,30000,0,20000
TDS Payable,Duties & Taxes,0,12000,0,10000
GST Payable,Duties & Taxes,0,18000,0,15000
Building,Fixed Assets,400000,0,400000,0
Plant & Machinery,Fixed Assets,200000,0,225000,0
Computer,Fixed Assets,50000,0,30000,0
Furniture,Fixed Assets,80000,0,80000,0
Sundry Debtor - Client A,Sundry Debtors,125000,0,100000,0
Sundry Debtor - Client B,Sundry Debtors,85000,0,90000,0
HDFC Bank - Current A/c,Bank Accounts,180000,0,130000,0
ICICI Bank - Current A/c,Bank Accounts,50000,0,20000,0
Cash on Hand,Cash-in-Hand,15000,0,12000,0
Raw Material Stock,Stock-in-Hand,120000,0,100000,0
Finished Goods,Stock-in-Hand,60000,0,60000,0
GST Input Credit,Duties & Taxes,40000,0,30000,0
Advance Tax,Loans & Advances (Asset),32000,0,8000,0
Prepaid Insurance,Loans & Advances (Asset),8000,0,6000,0
Security Deposit,Deposits (Asset),25000,0,25000,0
Sales - Products,Sales Accounts,0,2200000,0,1900000
Sales - Services,Sales Accounts,0,300000,0,300000
Interest Income,Indirect Income,0,15000,0,12000
Purchase - Raw Material,Purchase Accounts,1200000,0,1050000,0
Purchase - Traded Goods,Purchase Accounts,300000,0,250000,0
Salaries & Wages,Indirect Expenses,350000,0,330000,0
PF Contribution,Indirect Expenses,42000,0,39600,0
Staff Welfare,Indirect Expenses,18000,0,16000,0
Interest on Loan,Indirect Expenses,55000,0,65000,0
Bank Charges,Indirect Expenses,3500,0,3200,0
Depreciation,Indirect Expenses,95000,0,90000,0
Rent,Indirect Expenses,144000,0,135000,0
Power & Fuel,Indirect Expenses,72000,0,68000,0
Insurance,Indirect Expenses,22000,0,20000,0
Legal & Professional Fees,Indirect Expenses,35000,0,30000,0
Statutory Audit Fee,Indirect Expenses,25000,0,25000,0
Telephone,Indirect Expenses,15000,0,14000,0
Travelling,Indirect Expenses,28000,0,25000,0
Printing & Stationery,Indirect Expenses,8000,0,7000,0
Repair to Machinery,Indirect Expenses,22000,0,18000,0
Freight,Indirect Expenses,45000,0,40000,0
Advertisement,Indirect Expenses,12000,0,10000,0
Miscellaneous Expenses,Indirect Expenses,8500,0,8200,0
Current Tax,Duties & Taxes,65000,0,50000,0
"""
    r = c.post("/api/upload-tb/1",
               files={"file": ("tb.csv", io.BytesIO(sample_tb.encode()), "text/csv")},
               data={"replace": "true"})
    tb_res = r.json()
    check("TB upload", r.status_code == 200 and tb_res["rows_saved"] >= 45,
          f"rows={tb_res.get('rows_saved')}")

    # Auto-mapping
    r = c.post("/api/map/1/auto")
    m = r.json()
    check("Auto-map all", r.status_code == 200 and m["unmapped"] == 0,
          f"unmapped={m.get('unmapped')}")
    check("Keyword mapping used", m["mapped_by_keyword"] > 20,
          f"keyword={m.get('mapped_by_keyword')}")

    r = c.get("/api/map/1/summary")
    check("Mapping 100%", r.json()["completion_pct"] == 100)

    # Audit entries
    r = c.post("/api/audit/", json={
        "project_id": 1, "dr_coa_code": "PL-04-07-11",
        "cr_coa_code": "BS-EL-04-03-11", "amount": 25000,
        "narration": "Provision for professional fees", "status": "approved"
    })
    check("Create audit entry", r.status_code == 200)

    r = c.post("/api/audit/", json={
        "project_id": 1, "dr_coa_code": "PL-04-07-18",
        "cr_coa_code": "BS-AS-02-03-22", "amount": 15000,
        "narration": "Provision for doubtful debts", "status": "approved"
    })
    check("Create audit entry 2", r.status_code == 200)

    r = c.get("/api/audit/1/check")
    check("Audit balanced", r.json()["is_balanced"])

    # ========== SPRINT 2: ENGINE ==========
    print("\n--- SPRINT 2: Engine ---")

    # P&L
    r = c.get("/api/generate/1/pl")
    pl = r.json()
    check("P&L generated", r.status_code == 200 and pl["total_revenue_cy"] > 0,
          f"revenue={pl.get('total_revenue_cy')}")
    check("PAT computed", pl["pat_cy"] > 0, f"pat={pl.get('pat_cy')}")

    # BS
    r = c.get("/api/generate/1/bs")
    bs = r.json()
    check("BS generated", r.status_code == 200)
    check("BS has line items", len(bs["line_items"]) > 30)

    # Notes
    r = c.get("/api/generate/1/notes")
    notes = r.json()
    check("BS Notes: 20 generated", len(notes["bs_notes"]) == 20)
    check("PL Notes: 7 generated", len(notes["pl_notes"]) == 7)
    check("Note A (Share Capital) > 0", notes["bs_notes"]["A"]["total_cy"] > 0)
    check("Note Rev (Revenue) > 0", notes["pl_notes"]["Rev"]["total_cy"] > 0)
    check("Note OE has sub-items", len(notes["pl_notes"]["OE"]["items"]) == 27)

    # Cash Flow
    r = c.get("/api/generate/1/cashflow")
    cf = r.json()
    check("Cash Flow generated", r.status_code == 200)
    check("CF has line items", len(cf["line_items"]) > 30)

    # Ratios
    r = c.post("/api/generate/1/ratios")
    rs = r.json()
    check("11 ratios computed", len(rs["ratios"]) == 11)
    check("Variance flags work", "flagged_count" in rs)

    # EPS
    r = c.post("/api/generate/1/eps")
    eps = r.json()
    check("EPS computed", eps["basic_eps_cy"] > 0, f"eps={eps.get('basic_eps_cy')}")
    check("Face value correct", eps["face_value"] == 10)

    # All-in-one
    r = c.get("/api/generate/1/all")
    check("All-in-one endpoint", r.status_code == 200 and len(r.json()) == 6)

    # ========== SPRINT 3: EXPORT ==========
    print("\n--- SPRINT 3: Export ---")

    # Excel export
    r = c.get("/api/export/1/generate")
    check("Excel generated", r.status_code == 200 and r.json()["size_bytes"] > 50000,
          f"size={r.json().get('size_bytes')}")

    # Verify Excel content
    from openpyxl import load_workbook
    wb = load_workbook(r.json()["filepath"])
    check("Excel has 20 sheets", len(wb.sheetnames) == 20, f"sheets={len(wb.sheetnames)}")

    expected_sheets = ["Master", "TB", "Audit", "CoA", "Balance Sheet", "Profit & Loss",
                       "Notes-BS", "Notes-PL", "Cash Flow", "SOCIE", "TR Ageing", "TP Ageing",
                       "MSME", "PPE Schedule", "Related Party", "Acc Policies",
                       "Addl Disclosures", "EPS", "Ratios", "Contingent"]
    for s in expected_sheets:
        check(f"Sheet '{s}' exists", s in wb.sheetnames)

    # Ageing
    r = c.post("/api/ageing/", json={
        "project_id": 1, "ageing_type": "TR", "party_name": "Client A",
        "bucket_1": 50000, "bucket_2": 20000, "bucket_3": 10000, "period": "CY"
    })
    check("TR ageing entry", r.status_code == 200)

    r = c.get("/api/ageing/1/schedule/tr")
    check("TR schedule generated", r.status_code == 200)

    # Share Capital
    r = c.post("/api/share-events/", json={
        "project_id": 1, "event_type": "opening", "event_date": "2024-04-01",
        "no_of_shares": 10000, "face_value": 10, "period": "CY"
    })
    check("Share event created", r.status_code == 200)

    r = c.post("/api/shareholders/", json={
        "project_id": 1, "name": "Venkanna Setty E",
        "no_of_shares_cy": 8000, "no_of_shares_py": 7000, "is_promoter": True
    })
    check("Shareholder added", r.status_code == 200)

    r = c.get("/api/share-capital/1/full")
    check("Full Note A generated", r.status_code == 200 and "capital_structure" in r.json())

    r = c.get("/api/share-capital/1/promoters")
    check("Promoter holding with % change", r.status_code == 200 and len(r.json()) > 0)

    # ========== SPRINT 4: FRONTEND ==========
    print("\n--- SPRINT 4: Frontend ---")

    pages = [
        ("/", "Dashboard"),
        ("/project/1", "Project Detail"),
        ("/project/1/mapping", "Mapping Page"),
        ("/project/1/preview", "Preview Page"),
    ]
    for url, label in pages:
        r = c.get(url)
        check(f"Page: {label}", r.status_code == 200 and "<html" in r.text.lower())

    # ========== SPRINT 5: POLISH ==========
    print("\n--- SPRINT 5: Polish ---")

    # Related Party
    r = c.post("/api/related-parties/", json={
        "project_id": 1, "name": "Venkanna Setty E",
        "category": "KMP", "relationship": "Director"
    })
    check("Add related party", r.status_code == 200)

    r = c.post("/api/related-parties/", json={
        "project_id": 1, "name": "Setty Enterprises",
        "category": "Entity", "relationship": "Company in which Director is interested"
    })
    check("Add related entity", r.status_code == 200)

    r = c.post("/api/rp-transactions/", json={
        "project_id": 1, "party_id": 1,
        "transaction_type": "Remuneration", "cy_amount": 600000, "py_amount": 540000
    })
    check("RP transaction (Remuneration)", r.status_code == 200)

    r = c.post("/api/rp-transactions/", json={
        "project_id": 1, "party_id": 2,
        "transaction_type": "Purchase of Goods", "cy_amount": 150000, "py_amount": 120000
    })
    check("RP transaction (Purchase)", r.status_code == 200)

    r = c.get("/api/related-parties/1/disclosure")
    rp = r.json()
    check("RP disclosure generated", r.status_code == 200 and "matrix" in rp)
    check("RP matrix has transaction types", len(rp["matrix"]) == 16)
    check("RP Remuneration KMP CY=600000", rp["matrix"]["Remuneration"]["KMP"]["cy"] == 600000)

    # Accounting Policies
    r = c.get("/api/policies/1")
    check("Policies auto-seeded", r.status_code == 200 and len(r.json()) == 16)

    r = c.post("/api/policies/", json={
        "project_id": 1, "title": "Related Party Disclosures",
        "body": "Related party transactions are identified and disclosed as per AS-18."
    })
    check("Custom policy added", r.status_code == 200)

    r = c.get("/api/policies/1")
    check("Policies now 17", len(r.json()) == 17)

    # Additional Disclosures
    r = c.get("/api/disclosures/1")
    disc = r.json()
    check("Disclosures auto-seeded", r.status_code == 200 and len(disc) == 20)
    check("Disclosure A exists", "A" in disc and disc["A"]["title"] == "CIF VALUE OF IMPORTS")
    check("Disclosure Q (CSR) has 4 items", len(disc["Q"]["items"]) == 4)

    # ========== FINAL SUMMARY ==========
    print("\n--- ENDPOINT COUNT ---")
    api_count = len([rt for rt in app.routes if hasattr(rt, 'path') and '/api' in rt.path])
    web_count = 4
    print(f"  API endpoints: {api_count}")
    print(f"  Web pages: {web_count}")

    print(f"\n{'=' * 70}")
    print(f"RESULTS: {PASS} PASSED, {FAIL} FAILED out of {PASS + FAIL} checks")
    print(f"{'=' * 70}")

    if FAIL > 0:
        sys.exit(1)
