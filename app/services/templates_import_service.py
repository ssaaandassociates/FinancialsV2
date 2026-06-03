"""
Master-data and PPE template importers.

Symmetric with templates_service.py — reads the same Excel structure that the
download produces, and upserts records by their natural key:

  - Client Master    -> updates the Client row's fields by id
  - Directors        -> upsert by (client_id, name) lower-trimmed
  - Shareholders     -> upsert by (client_id, name) lower-trimmed
  - Custom CoA       -> upsert by (client_id, code)
  - Policies         -> upsert by (client_id, policy_number)
  - PPE              -> upsert by (project_id, coa_code)

For each sheet, we parse the header row to find columns by name so column-order
changes in future templates don't break the importer.
"""
from io import BytesIO
from typing import Any
from openpyxl import load_workbook
from sqlalchemy.orm import Session
from datetime import date, datetime

from app.models.client import (
    Client, Director, ClientShareholder, CustomCoACode, ClientPolicy
)
from app.models.supplementary import PPEScheduleEntry


# ---- helpers ----

def _coerce_bool(v: Any) -> bool:
    if isinstance(v, bool):
        return v
    if v is None:
        return False
    s = str(v).strip().lower()
    return s in ("true", "yes", "y", "1", "x", "✓")


def _coerce_num(v: Any) -> float | None:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _coerce_int(v: Any) -> int | None:
    n = _coerce_num(v)
    return int(n) if n is not None else None


def _coerce_date(v: Any) -> date | None:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%d %b %Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _header_index(ws, expected: list[str]) -> dict[str, int]:
    """
    Find the header row by scanning the first 8 rows for one that contains
    all of the expected header strings. Returns a map {header_name: col_index}.
    """
    expected_lc = {e.strip().lower() for e in expected}
    for row_idx in range(1, 9):
        cells = [c.value for c in ws[row_idx]]
        labels = {str(c).strip().lower(): i + 1 for i, c in enumerate(cells) if c}
        if expected_lc.issubset(set(labels.keys())):
            # Return original-case mapping limited to expected
            return {h: labels[h.strip().lower()] for h in expected}
    return {}


def _iter_data_rows(ws, header_row: int):
    """Yield rows after header_row, stopping at the first fully-blank row."""
    for row in ws.iter_rows(min_row=header_row + 1, values_only=False):
        if all(c.value is None or c.value == "" for c in row):
            break
        yield row


# ---- Client master ----

def _import_client_master(ws, client: Client) -> int:
    """
    Client Master sheet is a 2-column key/value layout:
        Field | Value | Notes
    Update the Client row in place from the values.
    """
    field_map = {
        "name": "name",
        "cin": "cin",
        "pan": "pan",
        "gstin": "gstin",
        "date of incorporation": "date_of_incorporation",
        "registered office": "registered_office",
        "principal activity": "principal_activity",
        "auditor name": "auditor_name",
        "auditor firm": "auditor_name",
        "auditor frn": "auditor_frn",
        "auditor frn (firm reg no.)": "auditor_frn",
        "auditor membership no": "auditor_membership_no",
        "auditor membership no.": "auditor_membership_no",
        "partner membership no.": "auditor_membership_no",
        "face value": "face_value",
        "authorised shares": "authorised_shares",
        "authorised capital": "authorised_capital",
        "subscribed shares": "subscribed_shares",
        "subscribed capital": "subscribed_capital",
        "paid-up shares": "paidup_shares",
        "paidup shares": "paidup_shares",
        "paid-up capital": "paidup_capital",
        "paidup capital": "paidup_capital",
    }
    updated = 0
    # Look at rows: col 1 = Field, col 2 = Value
    for row in ws.iter_rows(values_only=False):
        if not row or len(row) < 2:
            continue
        key_cell = row[0].value
        val_cell = row[1].value
        if key_cell is None:
            continue
        key = str(key_cell).strip().lower()
        attr = field_map.get(key)
        if not attr:
            continue
        if val_cell is None or val_cell == "":
            continue
        if attr == "date_of_incorporation":
            new = _coerce_date(val_cell)
        elif attr in ("face_value", "authorised_shares", "subscribed_shares",
                      "paidup_shares"):
            new = _coerce_int(val_cell)
        elif attr in ("authorised_capital", "subscribed_capital", "paidup_capital"):
            new = _coerce_num(val_cell)
        else:
            new = str(val_cell).strip()
        if new is not None and getattr(client, attr, None) != new:
            setattr(client, attr, new)
            updated += 1
    return updated


# ---- Directors ----

def _import_directors(db: Session, ws, client_id: int) -> tuple[int, int]:
    headers = ["Name", "DIN", "Designation", "Date of Appointment",
               "PAN", "Is KMP", "Signs Financials", "Is Active"]
    idx = _header_index(ws, headers)
    if not idx:
        return 0, 0
    header_row = next((r for r in range(1, 9)
                       if any(c.value and str(c.value).strip().lower() == "name"
                              for c in ws[r])), None)
    if header_row is None:
        return 0, 0

    existing = {d.name.strip().lower(): d
                for d in db.query(Director).filter(Director.client_id == client_id)}
    created = updated = 0

    for row in _iter_data_rows(ws, header_row):
        name = row[idx["Name"] - 1].value
        if not name:
            continue
        name_str = str(name).strip()
        key = name_str.lower()
        data = {
            "name": name_str,
            "din": str(row[idx["DIN"] - 1].value).strip() if row[idx["DIN"] - 1].value else None,
            "designation": str(row[idx["Designation"] - 1].value).strip() if row[idx["Designation"] - 1].value else "Director",
            "date_of_appointment": _coerce_date(row[idx["Date of Appointment"] - 1].value),
            "pan": str(row[idx["PAN"] - 1].value).strip() if row[idx["PAN"] - 1].value else None,
            "is_kmp": _coerce_bool(row[idx["Is KMP"] - 1].value),
            "signs_financials": _coerce_bool(row[idx["Signs Financials"] - 1].value),
            "is_active": _coerce_bool(row[idx["Is Active"] - 1].value) if row[idx["Is Active"] - 1].value not in (None, "") else True,
        }
        if key in existing:
            d = existing[key]
            for k, v in data.items():
                setattr(d, k, v)
            updated += 1
        else:
            db.add(Director(client_id=client_id, **data))
            created += 1
    db.commit()
    return created, updated


# ---- Shareholders ----

def _import_shareholders(db: Session, ws, client_id: int) -> tuple[int, int]:
    headers = ["Name", "No. of Shares (CY)", "No. of Shares (PY)",
               "Face Value", "% Holding (CY)", "% Holding (PY)",
               "Is Promoter", "Is Director", "DIN", "PAN"]
    idx = _header_index(ws, headers)
    if not idx:
        return 0, 0
    header_row = next((r for r in range(1, 9)
                       if any(c.value and str(c.value).strip().lower() == "name"
                              for c in ws[r])), None)
    if header_row is None:
        return 0, 0

    existing = {s.name.strip().lower(): s
                for s in db.query(ClientShareholder).filter(ClientShareholder.client_id == client_id)}
    created = updated = 0

    for row in _iter_data_rows(ws, header_row):
        name = row[idx["Name"] - 1].value
        if not name:
            continue
        name_str = str(name).strip()
        key = name_str.lower()
        data = {
            "name": name_str,
            "no_of_shares_cy": _coerce_int(row[idx["No. of Shares (CY)"] - 1].value),
            "no_of_shares_py": _coerce_int(row[idx["No. of Shares (PY)"] - 1].value),
            "face_value": _coerce_num(row[idx["Face Value"] - 1].value),
            "pct_holding_cy": _coerce_num(row[idx["% Holding (CY)"] - 1].value),
            "pct_holding_py": _coerce_num(row[idx["% Holding (PY)"] - 1].value),
            "is_promoter": _coerce_bool(row[idx["Is Promoter"] - 1].value),
            "is_director": _coerce_bool(row[idx["Is Director"] - 1].value),
            "din": str(row[idx["DIN"] - 1].value).strip() if row[idx["DIN"] - 1].value else None,
            "pan": str(row[idx["PAN"] - 1].value).strip() if row[idx["PAN"] - 1].value else None,
        }
        if key in existing:
            s = existing[key]
            for k, v in data.items():
                setattr(s, k, v)
            updated += 1
        else:
            db.add(ClientShareholder(client_id=client_id, **data))
            created += 1
    db.commit()
    return created, updated


# ---- Custom CoA ----

def _import_custom_coa(db: Session, ws, client_id: int) -> tuple[int, int]:
    headers = ["Code", "Particulars", "Parent Code",
               "Nature (Dr/Cr)", "FS Type (BS/PL)", "Note Ref"]
    idx = _header_index(ws, headers)
    if not idx:
        return 0, 0
    header_row = next((r for r in range(1, 9)
                       if any(c.value and str(c.value).strip().lower() == "code"
                              for c in ws[r])), None)
    if header_row is None:
        return 0, 0

    existing = {c.code.strip().upper(): c
                for c in db.query(CustomCoACode).filter(CustomCoACode.client_id == client_id)}
    created = updated = 0

    for row in _iter_data_rows(ws, header_row):
        code = row[idx["Code"] - 1].value
        if not code:
            continue
        code_str = str(code).strip().upper()
        data = {
            "code": code_str,
            "particulars": str(row[idx["Particulars"] - 1].value).strip() if row[idx["Particulars"] - 1].value else "",
            "parent_code": str(row[idx["Parent Code"] - 1].value).strip() if row[idx["Parent Code"] - 1].value else None,
            "nature": str(row[idx["Nature (Dr/Cr)"] - 1].value).strip() if row[idx["Nature (Dr/Cr)"] - 1].value else None,
            "fs_type": str(row[idx["FS Type (BS/PL)"] - 1].value).strip() if row[idx["FS Type (BS/PL)"] - 1].value else None,
            "note_ref": str(row[idx["Note Ref"] - 1].value).strip() if row[idx["Note Ref"] - 1].value else None,
        }
        if code_str in existing:
            c = existing[code_str]
            for k, v in data.items():
                setattr(c, k, v)
            updated += 1
        else:
            db.add(CustomCoACode(client_id=client_id, **data))
            created += 1
    db.commit()
    return created, updated


# ---- Policies ----

def _import_policies(db: Session, ws, client_id: int) -> tuple[int, int]:
    headers = ["Policy No.", "Title", "Body", "Is Active"]
    idx = _header_index(ws, headers)
    if not idx:
        return 0, 0
    header_row = next((r for r in range(1, 9)
                       if any(c.value and str(c.value).strip().lower() == "policy no."
                              for c in ws[r])), None)
    if header_row is None:
        return 0, 0

    existing = {p.policy_number: p
                for p in db.query(ClientPolicy).filter(ClientPolicy.client_id == client_id)}
    created = updated = 0

    for row in _iter_data_rows(ws, header_row):
        num = _coerce_int(row[idx["Policy No."] - 1].value)
        title = row[idx["Title"] - 1].value
        if not num or not title:
            continue
        data = {
            "policy_number": num,
            "title": str(title).strip(),
            "body": str(row[idx["Body"] - 1].value or "").strip(),
            "is_active": _coerce_bool(row[idx["Is Active"] - 1].value) if row[idx["Is Active"] - 1].value not in (None, "") else True,
        }
        if num in existing:
            p = existing[num]
            for k, v in data.items():
                setattr(p, k, v)
            updated += 1
        else:
            db.add(ClientPolicy(client_id=client_id, **data))
            created += 1
    db.commit()
    return created, updated


# ---- Public: Master data import ----

def import_master_data(db: Session, client_id: int, file_bytes: bytes) -> dict:
    """
    Imports a filled master-data workbook (blank template or export-current —
    both have the same sheet structure) and upserts everything for the given client.
    Returns per-sheet counts.
    """
    client = db.query(Client).filter(Client.id == client_id).first()
    if not client:
        raise ValueError(f"Client {client_id} not found")

    wb = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    sheets = {s.lower(): wb[s] for s in wb.sheetnames}

    result: dict = {"sheets_processed": [], "warnings": []}

    if "client master" in sheets:
        n = _import_client_master(sheets["client master"], client)
        db.commit()
        result["client_master_fields_updated"] = n
        result["sheets_processed"].append("Client Master")

    if "directors" in sheets:
        c, u = _import_directors(db, sheets["directors"], client_id)
        result["directors"] = {"created": c, "updated": u}
        result["sheets_processed"].append("Directors")

    if "shareholders" in sheets:
        c, u = _import_shareholders(db, sheets["shareholders"], client_id)
        result["shareholders"] = {"created": c, "updated": u}
        result["sheets_processed"].append("Shareholders")

    if "custom coa" in sheets:
        c, u = _import_custom_coa(db, sheets["custom coa"], client_id)
        result["custom_coa"] = {"created": c, "updated": u}
        result["sheets_processed"].append("Custom CoA")

    if "policies" in sheets:
        c, u = _import_policies(db, sheets["policies"], client_id)
        result["policies"] = {"created": c, "updated": u}
        result["sheets_processed"].append("Policies")

    if not result["sheets_processed"]:
        result["warnings"].append(
            "No recognised sheets found. Expected: Client Master, Directors, "
            "Shareholders, Custom CoA, Policies."
        )

    return result


# ---- Public: PPE import ----

def import_ppe(db: Session, project_id: int, file_bytes: bytes) -> dict:
    """
    Imports a PPE schedule. Matches on (project_id, coa_code) and updates the
    12 amount columns. Creates new rows if the CoA code doesn't exist yet for
    this project (rare — they auto-seed from BS-AS-01-* mappings, but allowed).
    """
    wb = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    # First sheet, regardless of name (templates_service uses "PPE Schedule")
    ws = wb[wb.sheetnames[0]]
    headers = [
        "CoA Code", "Asset Class", "Tangible/Intangible",
        "Gross Opening (CY)", "Gross Additions (CY)", "Gross Disposals (CY)",
        "Dep Opening (CY)", "Dep for Year (CY)", "Dep on Disposals (CY)",
        "Gross Opening (PY)", "Gross Additions (PY)", "Gross Disposals (PY)",
        "Dep Opening (PY)", "Dep for Year (PY)", "Dep on Disposals (PY)",
    ]
    idx = _header_index(ws, headers)
    if not idx:
        raise ValueError(
            "PPE sheet headers not found. Expected columns: " + ", ".join(headers)
        )
    header_row = next((r for r in range(1, 9)
                       if any(c.value and str(c.value).strip().lower() == "coa code"
                              for c in ws[r])), None)
    if header_row is None:
        raise ValueError("Could not locate header row")

    existing = {e.coa_code: e
                for e in db.query(PPEScheduleEntry).filter(PPEScheduleEntry.project_id == project_id)}
    updated = created = 0

    field_map = {
        "Gross Opening (CY)":   "gross_opening_cy",
        "Gross Additions (CY)": "gross_additions_cy",
        "Gross Disposals (CY)": "gross_disposals_cy",
        "Dep Opening (CY)":     "dep_opening_cy",
        "Dep for Year (CY)":    "dep_for_year_cy",
        "Dep on Disposals (CY)": "dep_on_disposals_cy",
        "Gross Opening (PY)":   "gross_opening_py",
        "Gross Additions (PY)": "gross_additions_py",
        "Gross Disposals (PY)": "gross_disposals_py",
        "Dep Opening (PY)":     "dep_opening_py",
        "Dep for Year (PY)":    "dep_for_year_py",
        "Dep on Disposals (PY)": "dep_on_disposals_py",
    }

    for row in _iter_data_rows(ws, header_row):
        code = row[idx["CoA Code"] - 1].value
        if not code:
            continue
        code_str = str(code).strip().upper()
        amounts = {f: _coerce_num(row[idx[h] - 1].value) or 0
                   for h, f in field_map.items()}
        if code_str in existing:
            e = existing[code_str]
            for f, v in amounts.items():
                setattr(e, f, v)
            updated += 1
        else:
            name = row[idx["Asset Class"] - 1].value
            db.add(PPEScheduleEntry(
                project_id=project_id,
                coa_code=code_str,
                particulars=str(name).strip() if name else code_str,
                **amounts,
            ))
            created += 1
    db.commit()
    return {"updated": updated, "created": created}
