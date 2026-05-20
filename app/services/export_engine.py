"""
Export Engine - generates the full 20-sheet Excel deliverable for a project.

Output sheets:
1. Master (company info, signatories, rounding)
2. TB (Trial Balance with audit adjustments)
3. Audit (proposed audit entries)
4. CoA (Chart of Accounts master)
5. Balance Sheet
6. Profit & Loss
7. Notes-BS (A through S)
8. Notes-PL (Rev, OI, Emp, Fin, Dep, OE, Tax)
9. Cash Flow
10. SOCIE (Statement of Changes in Equity - skeleton)
11. TR Ageing
12. TP Ageing
13. MSME Disclosure
14. PPE Schedule
15. Related Party (skeleton)
16. Accounting Policies (template)
17. Additional Disclosures (A through T)
18. EPS Computation
19. Ratios
20. Contingent Liabilities

Each sheet has:
- Company name header on row 1 (via =Master!B5)
- A4 portrait/landscape formatting
- Publication-ready styling
"""
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.models import Project, Client, TrialBalance, CoAMaster, AuditEntry
from app.services import (
    financial_engine, notes_engine, cashflow_engine,
    ratio_engine, eps_engine
)
from app.utils.excel_helpers import (
    setup_a4, set_cell, merge_title, header_row, section_row, set_column_widths,
    FONT_TITLE, FONT_SUBTITLE, FONT_HEADER, FONT_SECTION, FONT_ITEM,
    FONT_ITEM_LIGHT, FONT_INPUT, FONT_COMPUTED, FONT_PLAIN, FONT_NOTE,
    FONT_SMALL, FONT_RED_BOLD,
    FILL_HEADER, FILL_SECTION, FILL_INPUT, FILL_YELLOW, FILL_WHITE,
    FILL_LIGHT, FILL_GREEN, FILL_BLUE,
    ALIGN_LEFT, ALIGN_CENTER, ALIGN_RIGHT, ALIGN_WRAP,
    BORDER_ALL, BORDER_HAIR,
    NF_AMOUNT, NF_AMOUNT_2, NF_PCT, NF_RATIO,
    NAVY, GREEN
)
from app.config import OUTPUT_DIR

# Module-level context — set by export_full_excel before building sheets
_COMPANY_NAME = ""
_BS_HEADER_CY = "CY Amount"
_BS_HEADER_PY = "PY Amount"
_PL_HEADER_CY = "CY Amount"
_PL_HEADER_PY = "PY Amount"
_SIGNING = None  # SigningBlock dict or None


def _add_signing_block(ws, start_row, max_col=5):
    """Add signing block at bottom of a sheet."""
    if not _SIGNING:
        return start_row

    r = start_row + 2
    set_cell(ws, r, 1, "As per our report of even date attached", font=FONT_ITEM)
    r += 2
    set_cell(ws, r, 1, f"For {_SIGNING.get('auditor_firm', '')}", font=FONT_SECTION)
    set_cell(ws, r, 3, f"For and on behalf of Board of Directors of", font=FONT_SECTION)
    r += 1
    set_cell(ws, r, 1, "Chartered Accountants", font=FONT_PLAIN)
    set_cell(ws, r, 3, _COMPANY_NAME, font=FONT_SECTION)
    r += 1
    frn = _SIGNING.get('auditor_frn', '')
    if frn:
        set_cell(ws, r, 1, f"FRN: {frn}", font=FONT_PLAIN)
    r += 2
    set_cell(ws, r, 1, _SIGNING.get('partner_name', ''), font=FONT_SECTION)
    set_cell(ws, r, 3, _SIGNING.get('director1_name', ''), font=FONT_SECTION)
    set_cell(ws, r, 4 if max_col >= 5 else 3, _SIGNING.get('director2_name', ''), font=FONT_SECTION)
    r += 1
    set_cell(ws, r, 1, "Partner", font=FONT_PLAIN)
    set_cell(ws, r, 3, _SIGNING.get('director1_designation', 'Director'), font=FONT_PLAIN)
    set_cell(ws, r, 4 if max_col >= 5 else 3, _SIGNING.get('director2_designation', 'Director'), font=FONT_PLAIN)
    r += 1
    mno = _SIGNING.get('partner_membership_no', '')
    if mno:
        set_cell(ws, r, 1, f"M. No.: {mno}", font=FONT_PLAIN)
    d1din = _SIGNING.get('director1_din', '')
    d2din = _SIGNING.get('director2_din', '')
    if d1din:
        set_cell(ws, r, 3, f"DIN: {d1din}", font=FONT_PLAIN)
    if d2din:
        set_cell(ws, r, 4 if max_col >= 5 else 3, f"DIN: {d2din}", font=FONT_PLAIN)
    r += 2
    place = _SIGNING.get('place', '')
    sdate = _SIGNING.get('signing_date', '')
    set_cell(ws, r, 1, f"Place: {place}", font=FONT_PLAIN)
    set_cell(ws, r, 3, f"Place: {place}", font=FONT_PLAIN)
    r += 1
    set_cell(ws, r, 1, f"Date: {sdate}", font=FONT_PLAIN)
    set_cell(ws, r, 3, f"Date: {sdate}", font=FONT_PLAIN)
    r += 1
    udin = _SIGNING.get('partner_udin', '')
    if udin:
        set_cell(ws, r, 1, f"UDIN: {udin}", font=FONT_PLAIN)
    return r


# ============================================================
# SHEET BUILDERS
# ============================================================

def build_master_sheet(wb: Workbook, client: Client, project: Project):
    """Sheet 1: Master - company info and settings."""
    ws = wb.active
    ws.title = "Master"
    setup_a4(ws)
    ws.sheet_properties.tabColor = NAVY

    merge_title(ws, 1, "FINANCIAL STATEMENTS - TCE ENGINE", 7, FONT_TITLE)
    merge_title(ws, 2, "Schedule III Division I | Companies Act 2013", 7, FONT_SUBTITLE)

    r = 4
    section_row(ws, r, 1, "COMPANY INFORMATION", 7)

    fields = [
        ("Company Name", client.name or ""),
        ("CIN", client.cin or ""),
        ("Date of Incorporation", str(client.date_of_incorporation or "")),
        ("Registered Office", client.registered_office or ""),
        ("Financial Year", project.financial_year),
        ("BS Date (CY)", str(project.bs_date_cy or "")),
        ("BS Date (PY)", str(project.bs_date_py or "")),
        ("Rounding", project.rounding or "Rupees"),
        ("Auditor Name", client.auditor_name or ""),
        ("Auditor FRN", client.auditor_frn or ""),
        ("Tax Rate (%)", client.tax_rate or 0.2522),
        ("Authorised Capital (Rs.)", client.authorized_capital or 0),
        ("Paid-up Capital (Rs.)", client.paid_up_capital or 0),
        ("Face Value per Share (Rs.)", client.face_value or 10),
    ]
    for i, (label, value) in enumerate(fields):
        r2 = 5 + i
        set_cell(ws, r2, 1, label, font=FONT_ITEM, border=BORDER_ALL, align=ALIGN_LEFT)
        ws.merge_cells(start_row=r2, start_column=2, end_row=r2, end_column=4)
        set_cell(ws, r2, 2, value, font=FONT_INPUT, fill=FILL_INPUT,
                 border=BORDER_ALL, align=ALIGN_LEFT)
        for c in [3, 4]:
            ws.cell(row=r2, column=c).border = BORDER_ALL
            ws.cell(row=r2, column=c).fill = FILL_INPUT

    set_column_widths(ws, {'A': 28, 'B': 28, 'C': 18, 'D': 18, 'E': 16})
    ws.freeze_panes = 'A4'


def build_tb_sheet(wb: Workbook, db: Session, project_id: int):
    """Sheet 2: Trial Balance with Audit Adjustments."""
    ws = wb.create_sheet("TB")
    setup_a4(ws, 'landscape')
    ws.sheet_properties.tabColor = "2196F3"

    merge_title(ws, 1, _COMPANY_NAME, 14, FONT_TITLE)
    merge_title(ws, 2, "Trial Balance with Audit Adjustments", 14, FONT_SUBTITLE)

    r = 4
    headers = [
        (1, "S.No."), (2, "CoA Code"), (3, "Ledger Name"), (4, "Tally Group"),
        (5, "CY Dr (Raw)"), (6, "CY Cr (Raw)"),
        (7, "Adj Dr"), (8, "Adj Cr"),
        (9, "CY Dr (Final)"), (10, "CY Cr (Final)"), (11, "CY Net"),
        (12, "PY Dr"), (13, "PY Cr"), (14, "PY Net"),
    ]
    header_row(ws, r, headers)

    # Get TB rows
    tb_rows = db.query(TrialBalance).filter(
        TrialBalance.project_id == project_id
    ).all()

    # Get audit adjustments summed by code
    from app.services.audit_service import get_audit_adjustments_by_code
    adjustments = get_audit_adjustments_by_code(db, project_id, approved_only=True)

    for i, row in enumerate(tb_rows):
        r2 = r + 1 + i
        adj = adjustments.get(row.coa_code, {"debit": 0, "credit": 0})
        adj_dr = adj["debit"]
        adj_cr = adj["credit"]
        final_dr = (row.cy_debit or 0) + adj_dr
        final_cr = (row.cy_credit or 0) + adj_cr
        cy_net_final = final_dr - final_cr

        set_cell(ws, r2, 1, i + 1, font=FONT_ITEM, align=ALIGN_CENTER, border=BORDER_HAIR)
        set_cell(ws, r2, 2, row.coa_code or "", font=FONT_PLAIN, border=BORDER_HAIR)
        set_cell(ws, r2, 3, row.ledger_name, font=FONT_ITEM, border=BORDER_HAIR, align=ALIGN_LEFT)
        set_cell(ws, r2, 4, row.tally_group, font=FONT_ITEM_LIGHT, border=BORDER_HAIR, align=ALIGN_LEFT)
        set_cell(ws, r2, 5, row.cy_debit or 0, font=FONT_PLAIN, border=BORDER_HAIR, fmt=NF_AMOUNT)
        set_cell(ws, r2, 6, row.cy_credit or 0, font=FONT_PLAIN, border=BORDER_HAIR, fmt=NF_AMOUNT)
        set_cell(ws, r2, 7, adj_dr, font=FONT_COMPUTED, border=BORDER_HAIR, fmt=NF_AMOUNT)
        set_cell(ws, r2, 8, adj_cr, font=FONT_COMPUTED, border=BORDER_HAIR, fmt=NF_AMOUNT)
        set_cell(ws, r2, 9, final_dr, font=FONT_PLAIN, border=BORDER_HAIR, fmt=NF_AMOUNT)
        set_cell(ws, r2, 10, final_cr, font=FONT_PLAIN, border=BORDER_HAIR, fmt=NF_AMOUNT)
        set_cell(ws, r2, 11, cy_net_final, font=FONT_PLAIN, border=BORDER_HAIR, fmt=NF_AMOUNT)
        set_cell(ws, r2, 12, row.py_debit or 0, font=FONT_PLAIN, border=BORDER_HAIR, fmt=NF_AMOUNT)
        set_cell(ws, r2, 13, row.py_credit or 0, font=FONT_PLAIN, border=BORDER_HAIR, fmt=NF_AMOUNT)
        set_cell(ws, r2, 14, (row.py_debit or 0) - (row.py_credit or 0),
                 font=FONT_PLAIN, border=BORDER_HAIR, fmt=NF_AMOUNT)

    # Total row
    tr = r + 1 + len(tb_rows)
    set_cell(ws, tr, 3, "TOTAL", font=FONT_SECTION, border=BORDER_ALL)
    for c in range(5, 15):
        col = get_column_letter(c)
        set_cell(ws, tr, c, f"=SUM({col}{r+1}:{col}{tr-1})",
                 font=FONT_SECTION, border=BORDER_ALL, fmt=NF_AMOUNT)

    # Check row
    cr = tr + 1
    set_cell(ws, cr, 3, "CHECK (Dr-Cr)", font=FONT_RED_BOLD)
    set_cell(ws, cr, 9, f"=I{tr}-J{tr}", font=FONT_RED_BOLD, fmt=NF_AMOUNT)
    set_cell(ws, cr, 12, f"=L{tr}-M{tr}", font=FONT_RED_BOLD, fmt=NF_AMOUNT)

    set_column_widths(ws, {
        'A': 5, 'B': 18, 'C': 32, 'D': 18,
        'E': 14, 'F': 14, 'G': 14, 'H': 14,
        'I': 14, 'J': 14, 'K': 14, 'L': 14, 'M': 14, 'N': 14,
    })
    ws.auto_filter.ref = f"A4:N{tr}"
    ws.freeze_panes = 'A5'


def build_audit_sheet(wb: Workbook, db: Session, project_id: int):
    """Sheet 3: Audit Proposed Entries."""
    ws = wb.create_sheet("Audit")
    setup_a4(ws, 'landscape')
    ws.sheet_properties.tabColor = "F44336"

    merge_title(ws, 1, _COMPANY_NAME, 10, FONT_TITLE)
    merge_title(ws, 2, "Audit Proposed Adjustment Entries", 10, FONT_SUBTITLE)

    r = 4
    headers = [
        (1, "Entry No."), (2, "Date"), (3, "Narration"),
        (4, "Dr CoA Code"), (5, "Cr CoA Code"),
        (6, "Dr Amount"), (7, "Dr Ledger Name"),
        (8, "Cr Amount"), (9, "Cr Ledger Name"),
        (10, "Status"),
    ]
    header_row(ws, r, headers)

    entries = db.query(AuditEntry).filter(
        AuditEntry.project_id == project_id
    ).order_by(AuditEntry.entry_no).all()

    for i, e in enumerate(entries):
        r2 = r + 1 + i
        dr_coa = db.query(CoAMaster).filter(CoAMaster.code == e.dr_coa_code).first()
        cr_coa = db.query(CoAMaster).filter(CoAMaster.code == e.cr_coa_code).first()

        set_cell(ws, r2, 1, e.entry_no, font=FONT_ITEM, align=ALIGN_CENTER, border=BORDER_HAIR)
        set_cell(ws, r2, 2, str(e.date) if e.date else "", font=FONT_ITEM, border=BORDER_HAIR, align=ALIGN_CENTER)
        set_cell(ws, r2, 3, e.narration or "", font=FONT_ITEM, border=BORDER_HAIR, align=ALIGN_LEFT)
        set_cell(ws, r2, 4, e.dr_coa_code, font=FONT_PLAIN, border=BORDER_HAIR)
        set_cell(ws, r2, 5, e.cr_coa_code, font=FONT_PLAIN, border=BORDER_HAIR)
        set_cell(ws, r2, 6, e.amount, font=FONT_PLAIN, border=BORDER_HAIR, fmt=NF_AMOUNT)
        set_cell(ws, r2, 7, dr_coa.particulars if dr_coa else "", font=FONT_ITEM_LIGHT, border=BORDER_HAIR)
        set_cell(ws, r2, 8, e.amount, font=FONT_PLAIN, border=BORDER_HAIR, fmt=NF_AMOUNT)
        set_cell(ws, r2, 9, cr_coa.particulars if cr_coa else "", font=FONT_ITEM_LIGHT, border=BORDER_HAIR)
        set_cell(ws, r2, 10, e.status, font=FONT_ITEM, border=BORDER_HAIR, align=ALIGN_CENTER)

    # Total row
    if entries:
        tr = r + 1 + len(entries)
        set_cell(ws, tr, 5, "TOTAL", font=FONT_SECTION, border=BORDER_ALL)
        set_cell(ws, tr, 6, f"=SUM(F{r+1}:F{tr-1})", font=FONT_SECTION, border=BORDER_ALL, fmt=NF_AMOUNT)
        set_cell(ws, tr, 8, f"=SUM(H{r+1}:H{tr-1})", font=FONT_SECTION, border=BORDER_ALL, fmt=NF_AMOUNT)
        cr2 = tr + 1
        set_cell(ws, cr2, 5, "CHECK (Dr=Cr)", font=FONT_RED_BOLD)
        set_cell(ws, cr2, 6, f"=F{tr}-H{tr}", font=FONT_RED_BOLD, fmt=NF_AMOUNT)

    set_column_widths(ws, {
        'A': 8, 'B': 12, 'C': 40, 'D': 18, 'E': 18,
        'F': 16, 'G': 25, 'H': 16, 'I': 25, 'J': 12,
    })


def build_coa_sheet(wb: Workbook, db: Session):
    """Sheet 4: Chart of Accounts master."""
    ws = wb.create_sheet("CoA")
    setup_a4(ws, 'landscape')
    ws.sheet_properties.tabColor = GREEN

    merge_title(ws, 1, "Chart of Accounts - Schedule III Division I", 9, FONT_TITLE)

    r = 3
    headers = [
        (1, "Code"), (2, "Lvl"), (3, "Particulars"), (4, "Sch III Ref"),
        (5, "Dr/Cr"), (6, "Stmt"), (7, "Note"), (8, "Tally Group"), (9, "Remarks"),
    ]
    header_row(ws, r, headers)

    codes = db.query(CoAMaster).order_by(CoAMaster.code).all()
    for i, c in enumerate(codes):
        r2 = r + 1 + i
        set_cell(ws, r2, 1, c.code, font=FONT_SMALL, border=BORDER_HAIR)
        set_cell(ws, r2, 2, c.level, font=FONT_ITEM_LIGHT, border=BORDER_HAIR, align=ALIGN_CENTER)
        set_cell(ws, r2, 3, c.particulars, font=FONT_ITEM_LIGHT, border=BORDER_HAIR, align=ALIGN_LEFT)
        set_cell(ws, r2, 4, c.schedule_ref or "", font=FONT_SMALL, border=BORDER_HAIR, align=ALIGN_CENTER)
        set_cell(ws, r2, 5, c.nature or "", font=FONT_SMALL, border=BORDER_HAIR, align=ALIGN_CENTER)
        set_cell(ws, r2, 6, c.fs_type or "", font=FONT_SMALL, border=BORDER_HAIR, align=ALIGN_CENTER)
        set_cell(ws, r2, 7, c.note_ref or "", font=FONT_SMALL, border=BORDER_HAIR, align=ALIGN_CENTER)
        set_cell(ws, r2, 8, c.tally_group or "", font=FONT_SMALL, border=BORDER_HAIR)
        set_cell(ws, r2, 9, c.remarks or "", font=FONT_SMALL, border=BORDER_HAIR)

    set_column_widths(ws, {
        'A': 22, 'B': 5, 'C': 38, 'D': 12, 'E': 8,
        'F': 6, 'G': 8, 'H': 20, 'I': 18,
    })
    ws.auto_filter.ref = f"A3:I{r + 1 + len(codes)}"
    ws.freeze_panes = 'A4'


def build_balance_sheet(wb: Workbook, db: Session, project_id: int):
    """Sheet 5: Balance Sheet face."""
    ws = wb.create_sheet("Balance Sheet")
    setup_a4(ws)
    ws.sheet_properties.tabColor = NAVY

    merge_title(ws, 1, _COMPANY_NAME, 5, FONT_TITLE)
    merge_title(ws, 2, f"Balance Sheet {_BS_HEADER_CY}", 5, FONT_SUBTITLE)
    set_cell(ws, 3, 5, "(Amount in Indian Rupees)", font=FONT_NOTE, align=ALIGN_RIGHT)

    r = 4
    header_row(ws, r, [(1, ""), (2, "Particulars"), (3, "Note"),
                       (4, _BS_HEADER_CY), (5, _BS_HEADER_PY)])

    bs = financial_engine.generate_bs(db, project_id)

    for i, (sno, text, note_ref, cy, py, level, is_total) in enumerate(bs["line_items"]):
        r2 = r + 1 + i
        is_major_total = is_total and level == 1
        is_sub_total = is_total and level == 2
        is_section_header = level == 1 and not is_total

        f = FONT_SECTION if (is_total or level <= 2) else FONT_ITEM
        fl = FILL_SECTION if is_major_total else (FILL_BLUE if is_sub_total else FILL_WHITE)
        if is_section_header:
            fl = FILL_SECTION
            f = FONT_SECTION

        set_cell(ws, r2, 1, sno, font=f, fill=fl, border=BORDER_ALL, align=ALIGN_CENTER)
        set_cell(ws, r2, 2, text, font=f, fill=fl, border=BORDER_ALL, align=ALIGN_LEFT)
        set_cell(ws, r2, 3, note_ref, font=FONT_NOTE, fill=fl, border=BORDER_ALL, align=ALIGN_CENTER)
        if cy is not None:
            set_cell(ws, r2, 4, round(cy, 2), font=f, fill=fl, border=BORDER_ALL,
                     fmt=NF_AMOUNT, align=ALIGN_RIGHT)
        else:
            set_cell(ws, r2, 4, None, font=f, fill=fl, border=BORDER_ALL)
        if py is not None:
            set_cell(ws, r2, 5, round(py, 2), font=f, fill=fl, border=BORDER_ALL,
                     fmt=NF_AMOUNT, align=ALIGN_RIGHT)
        else:
            set_cell(ws, r2, 5, None, font=f, fill=fl, border=BORDER_ALL)

    # Balance check footer
    bs_end = r + 1 + len(bs["line_items"])
    set_cell(ws, bs_end, 2, "Difference (should be 0)", font=FONT_RED_BOLD)
    set_cell(ws, bs_end, 4, round(bs["difference_cy"], 2), font=FONT_RED_BOLD, fmt=NF_AMOUNT)
    set_cell(ws, bs_end, 5, round(bs["difference_py"], 2), font=FONT_RED_BOLD, fmt=NF_AMOUNT)

    set_cell(ws, bs_end + 1, 2, "See accompanying notes to the Financial Statements", font=FONT_NOTE)

    _add_signing_block(ws, bs_end + 2)

    set_column_widths(ws, {'A': 5, 'B': 45, 'C': 6, 'D': 20, 'E': 20})
    ws.freeze_panes = 'A5'


def build_pl_sheet(wb: Workbook, db: Session, project_id: int):
    """Sheet 6: Profit & Loss face."""
    ws = wb.create_sheet("Profit & Loss")
    setup_a4(ws)
    ws.sheet_properties.tabColor = "E65100"

    merge_title(ws, 1, _COMPANY_NAME, 5, FONT_TITLE)
    merge_title(ws, 2, f"Statement of Profit and Loss {_PL_HEADER_CY}", 5, FONT_SUBTITLE)
    set_cell(ws, 3, 5, "(Amount in Indian Rupees)", font=FONT_NOTE, align=ALIGN_RIGHT)

    r = 4
    header_row(ws, r, [(1, ""), (2, "Particulars"), (3, "Note"),
                       (4, _PL_HEADER_CY), (5, _PL_HEADER_PY)])

    pl = financial_engine.generate_pl(db, project_id)

    for i, item in enumerate(pl["line_items"]):
        # Item structure: (sno, text, note_ref, cy, py, level, is_total)
        sno, text, note_ref, cy, py, level, is_total = item
        r2 = r + 1 + i
        f = FONT_SECTION if is_total else FONT_ITEM
        fl = FILL_SECTION if is_total else FILL_WHITE

        set_cell(ws, r2, 1, sno, font=f, fill=fl, border=BORDER_ALL, align=ALIGN_CENTER)
        set_cell(ws, r2, 2, text, font=f, fill=fl, border=BORDER_ALL, align=ALIGN_LEFT)
        set_cell(ws, r2, 3, note_ref, font=FONT_NOTE, fill=fl, border=BORDER_ALL, align=ALIGN_CENTER)
        if cy is not None:
            set_cell(ws, r2, 4, round(cy, 2), font=f, fill=fl, border=BORDER_ALL,
                     fmt=NF_AMOUNT, align=ALIGN_RIGHT)
        else:
            set_cell(ws, r2, 4, None, font=f, fill=fl, border=BORDER_ALL)
        if py is not None:
            set_cell(ws, r2, 5, round(py, 2), font=f, fill=fl, border=BORDER_ALL,
                     fmt=NF_AMOUNT, align=ALIGN_RIGHT)
        else:
            set_cell(ws, r2, 5, None, font=f, fill=fl, border=BORDER_ALL)

    _add_signing_block(ws, r + 1 + len(pl["line_items"]))

    set_column_widths(ws, {'A': 6, 'B': 50, 'C': 6, 'D': 20, 'E': 20})
    ws.freeze_panes = 'A5'


def build_notes_bs_sheet(wb: Workbook, db: Session, project_id: int):
    """Sheet 7: Notes for Balance Sheet (A through S)."""
    ws = wb.create_sheet("Notes-BS")
    setup_a4(ws)
    ws.sheet_properties.tabColor = "1565C0"

    merge_title(ws, 1, _COMPANY_NAME, 5, FONT_TITLE)
    merge_title(ws, 2, "Notes to the Balance Sheet", 5, FONT_SUBTITLE)
    set_cell(ws, 3, 5, "(Amount in Indian Rupees)", font=FONT_NOTE, align=ALIGN_RIGHT)

    notes = notes_engine.generate_all_notes(db, project_id)
    bs_notes = notes["bs_notes"]

    note_titles = {
        "A": "Share Capital", "B": "Reserves and Surplus", "C": "Long-Term Borrowings",
        "D": "Other Long-Term Liabilities", "E": "Long-Term Provisions",
        "F": "Short-Term Borrowings", "FA": "Trade Payables (MSME Disclosure)",
        "G": "Other Current Liabilities", "H": "Short-Term Provisions",
        "I": "Tangible Assets", "J": "Intangible Assets", "K": "Non-Current Investments",
        "L": "Long-Term Loans & Advances", "M": "Other Non-Current Assets",
        "N": "Current Investments", "O": "Inventories", "P": "Trade Receivables",
        "Q": "Cash and Cash Equivalents", "R": "Short-Term Loans & Advances",
        "S": "Other Current Assets",
    }

    r = 5
    for key in ["A", "B", "C", "D", "E", "F", "FA", "G", "H", "I", "J", "K",
                "L", "M", "N", "O", "P", "Q", "R", "S"]:
        note = bs_notes.get(key)
        if not note:
            continue

        # Note header
        section_row(ws, r, 1, f"Note {key}: {note_titles.get(key, '')}", 5)
        r += 1

        # Column headers
        header_row(ws, r, [(1, ""), (2, "Particulars"), (3, ""),
                           (4, _BS_HEADER_CY), (5, _BS_HEADER_PY)])
        r += 1

        # Note items
        for item in note["items"]:
            sno = item.get("sno", "")
            set_cell(ws, r, 1, sno, font=FONT_ITEM, fill=FILL_WHITE,
                     border=BORDER_ALL, align=ALIGN_CENTER)
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
            set_cell(ws, r, 2, item["particulars"], font=FONT_ITEM,
                     fill=FILL_WHITE, border=BORDER_ALL, align=ALIGN_LEFT)
            set_cell(ws, r, 3, None, font=FONT_ITEM, fill=FILL_WHITE, border=BORDER_ALL)
            set_cell(ws, r, 4, round(item["cy_amount"], 2), font=FONT_ITEM,
                     fill=FILL_WHITE, border=BORDER_ALL, fmt=NF_AMOUNT, align=ALIGN_RIGHT)
            set_cell(ws, r, 5, round(item["py_amount"], 2), font=FONT_ITEM,
                     fill=FILL_WHITE, border=BORDER_ALL, fmt=NF_AMOUNT, align=ALIGN_RIGHT)
            r += 1

        # Total row
        set_cell(ws, r, 1, "", font=FONT_SECTION, fill=FILL_SECTION, border=BORDER_ALL)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        set_cell(ws, r, 2, f"Total Note {key}", font=FONT_SECTION,
                 fill=FILL_SECTION, border=BORDER_ALL, align=ALIGN_LEFT)
        set_cell(ws, r, 3, None, font=FONT_SECTION, fill=FILL_SECTION, border=BORDER_ALL)
        set_cell(ws, r, 4, round(note["total_cy"], 2), font=FONT_SECTION,
                 fill=FILL_SECTION, border=BORDER_ALL, fmt=NF_AMOUNT, align=ALIGN_RIGHT)
        set_cell(ws, r, 5, round(note["total_py"], 2), font=FONT_SECTION,
                 fill=FILL_SECTION, border=BORDER_ALL, fmt=NF_AMOUNT, align=ALIGN_RIGHT)
        r += 2  # Blank spacer row

    set_column_widths(ws, {'A': 5, 'B': 35, 'C': 10, 'D': 20, 'E': 20})
    ws.freeze_panes = 'A5'


def build_notes_pl_sheet(wb: Workbook, db: Session, project_id: int):
    """Sheet 8: Notes for Profit & Loss."""
    ws = wb.create_sheet("Notes-PL")
    setup_a4(ws)
    ws.sheet_properties.tabColor = "E65100"

    merge_title(ws, 1, _COMPANY_NAME, 5, FONT_TITLE)
    merge_title(ws, 2, "Notes to the Profit & Loss Statement", 5, FONT_SUBTITLE)
    set_cell(ws, 3, 5, "(Amount in Indian Rupees)", font=FONT_NOTE, align=ALIGN_RIGHT)

    notes = notes_engine.generate_all_notes(db, project_id)
    pl_notes = notes["pl_notes"]

    note_titles = {
        "Rev": "Revenue from Operations", "OI": "Other Income",
        "Emp": "Employee Benefits Expense", "Fin": "Finance Costs",
        "Dep": "Depreciation and Amortization", "OE": "Other Expenses",
        "Tax": "Tax Expense",
    }

    r = 5
    for key in ["Rev", "OI", "Emp", "Fin", "Dep", "OE", "Tax"]:
        note = pl_notes.get(key)
        if not note:
            continue

        section_row(ws, r, 1, f"Note {key}: {note_titles.get(key, '')}", 5)
        r += 1

        header_row(ws, r, [(1, ""), (2, "Particulars"), (3, ""),
                           (4, _PL_HEADER_CY), (5, _PL_HEADER_PY)])
        r += 1

        for item in note["items"]:
            sno = item.get("sno", "")
            set_cell(ws, r, 1, sno, font=FONT_ITEM, fill=FILL_WHITE,
                     border=BORDER_ALL, align=ALIGN_CENTER)
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
            set_cell(ws, r, 2, item["particulars"], font=FONT_ITEM,
                     fill=FILL_WHITE, border=BORDER_ALL, align=ALIGN_LEFT)
            set_cell(ws, r, 3, None, font=FONT_ITEM, fill=FILL_WHITE, border=BORDER_ALL)
            set_cell(ws, r, 4, round(item["cy_amount"], 2), font=FONT_ITEM,
                     fill=FILL_WHITE, border=BORDER_ALL, fmt=NF_AMOUNT, align=ALIGN_RIGHT)
            set_cell(ws, r, 5, round(item["py_amount"], 2), font=FONT_ITEM,
                     fill=FILL_WHITE, border=BORDER_ALL, fmt=NF_AMOUNT, align=ALIGN_RIGHT)
            r += 1

        set_cell(ws, r, 1, "", font=FONT_SECTION, fill=FILL_SECTION, border=BORDER_ALL)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        set_cell(ws, r, 2, f"Total Note {key}", font=FONT_SECTION,
                 fill=FILL_SECTION, border=BORDER_ALL, align=ALIGN_LEFT)
        set_cell(ws, r, 3, None, font=FONT_SECTION, fill=FILL_SECTION, border=BORDER_ALL)
        set_cell(ws, r, 4, round(note["total_cy"], 2), font=FONT_SECTION,
                 fill=FILL_SECTION, border=BORDER_ALL, fmt=NF_AMOUNT, align=ALIGN_RIGHT)
        set_cell(ws, r, 5, round(note["total_py"], 2), font=FONT_SECTION,
                 fill=FILL_SECTION, border=BORDER_ALL, fmt=NF_AMOUNT, align=ALIGN_RIGHT)
        r += 2

    set_column_widths(ws, {'A': 5, 'B': 35, 'C': 10, 'D': 20, 'E': 20})
def build_cashflow_sheet(wb, db: Session, project_id: int):
    """Sheet 9: Cash Flow Statement (Indirect Method)."""
    ws = wb.create_sheet("Cash Flow")
    setup_a4(ws)
    ws.sheet_properties.tabColor = "4CAF50"

    merge_title(ws, 1, _COMPANY_NAME, 5, FONT_TITLE)
    merge_title(ws, 2, f"Cash Flow Statement (Indirect Method) {_PL_HEADER_CY}", 5, FONT_SUBTITLE)
    set_cell(ws, 3, 5, "(Amount in Indian Rupees)", font=FONT_NOTE, align=ALIGN_RIGHT)

    r = 4
    header_row(ws, r, [(1, ""), (2, "Particulars"), (3, ""),
                       (4, _PL_HEADER_CY), (5, _PL_HEADER_PY)])

    cf = cashflow_engine.generate_cashflow(db, project_id)

    for i, item in enumerate(cf["line_items"]):
        ref, text, cy, py, is_total, level = item
        r2 = r + 1 + i
        if not text:
            continue

        f = FONT_SECTION if is_total else FONT_ITEM
        fl = FILL_SECTION if (is_total and level == 1) else (
            FILL_BLUE if is_total else FILL_WHITE
        )

        set_cell(ws, r2, 1, ref, font=f, fill=fl, border=BORDER_ALL, align=ALIGN_CENTER)
        ws.merge_cells(start_row=r2, start_column=2, end_row=r2, end_column=3)
        set_cell(ws, r2, 2, text, font=f, fill=fl, border=BORDER_ALL, align=ALIGN_LEFT)
        set_cell(ws, r2, 3, None, font=f, fill=fl, border=BORDER_ALL)
        if cy is not None:
            set_cell(ws, r2, 4, round(cy, 2), font=f, fill=fl, border=BORDER_ALL,
                     fmt=NF_AMOUNT, align=ALIGN_RIGHT)
        else:
            set_cell(ws, r2, 4, None, font=f, fill=fl, border=BORDER_ALL)
        if py is not None:
            set_cell(ws, r2, 5, round(py, 2), font=f, fill=fl, border=BORDER_ALL,
                     fmt=NF_AMOUNT, align=ALIGN_RIGHT)
        else:
            set_cell(ws, r2, 5, None, font=f, fill=fl, border=BORDER_ALL)

    set_column_widths(ws, {'A': 4, 'B': 48, 'C': 2, 'D': 20, 'E': 20})


def build_socie_sheet(wb):
    """Sheet 10: Statement of Changes in Equity (skeleton for manual input)."""
    ws = wb.create_sheet("SOCIE")
    setup_a4(ws)
    ws.sheet_properties.tabColor = "9C27B0"

    merge_title(ws, 1, _COMPANY_NAME, 7, FONT_TITLE)
    merge_title(ws, 2, "Statement of Changes in Equity", 7, FONT_SUBTITLE)

    r = 4
    section_row(ws, r, 1, "A. EQUITY SHARE CAPITAL", 7)
    header_row(ws, r + 1, [(1, "Opening Balance"), (3, "Changes During Year"),
                           (5, "Closing Balance")])
    for c in [1, 3, 5]:
        set_cell(ws, r + 2, c, None, font=FONT_INPUT, fill=FILL_INPUT,
                 border=BORDER_ALL, fmt=NF_AMOUNT)

    r2 = r + 5
    section_row(ws, r2, 1, "B. OTHER EQUITY", 7)
    header_row(ws, r2 + 1, [
        (1, "Particulars"), (2, "Capital Res"), (3, "Sec Premium"),
        (4, "General Res"), (5, "Retained Earn"), (6, "Other"), (7, "Total")
    ])

    rows = ["Opening Balance", "Profit for Year", "Other Comprehensive Income",
            "Dividends", "Transfer to Reserves", "Other Changes", "Closing Balance"]
    for i, label in enumerate(rows):
        r3 = r2 + 2 + i
        is_total = "Closing" in label or "Opening" in label
        set_cell(ws, r3, 1, label,
                 font=FONT_SECTION if is_total else FONT_ITEM,
                 border=BORDER_ALL)
        for c in range(2, 8):
            set_cell(ws, r3, c, None, font=FONT_INPUT, fill=FILL_INPUT,
                     border=BORDER_ALL, fmt=NF_AMOUNT)

    set_column_widths(ws, {'A': 25, 'B': 16, 'C': 16, 'D': 16, 'E': 16, 'F': 16, 'G': 16})


def build_tr_ageing_sheet(wb, db: Session, project_id: int):
    """Sheet 11: Trade Receivables Ageing (22 buckets)."""
    ws = wb.create_sheet("TR Ageing")
    setup_a4(ws, 'landscape')
    ws.sheet_properties.tabColor = "00897B"

    merge_title(ws, 1, _COMPANY_NAME, 7, FONT_TITLE)
    merge_title(ws, 2, "Trade Receivables - Ageing Schedule", 7, FONT_SUBTITLE)

    from app.services.financial_engine import get_adjusted_balances
    balances = get_adjusted_balances(db, project_id)

    def get_bucket(code):
        cy = balances.get(code, {}).get("cy_net", 0)
        py = balances.get(code, {}).get("py_net", 0)
        return cy, py

    for label, start_row in [("Current Reporting Period", 5), ("Previous Reporting Period", 19)]:
        set_cell(ws, start_row, 1, f"Figures for the {label}", font=FONT_SECTION)
        r2 = start_row + 1
        header_row(ws, r2, [
            (1, "Particulars"), (2, "<6M"), (3, "6M-1Y"),
            (4, "1-2Y"), (5, "2-3Y"), (6, ">3Y"), (7, "Total")
        ])

        # Undisputed Good: codes 01-05
        # Undisputed Doubtful: 06-10
        # Disputed Good: 11-15
        # Disputed Doubtful: 16-20
        rows_data = [
            ("Undisputed Trade Receivables", None, True),
            ("  Considered Good",
             [("BS-AS-02-03-01"), ("BS-AS-02-03-02"), ("BS-AS-02-03-03"),
              ("BS-AS-02-03-04"), ("BS-AS-02-03-05")], False),
            ("  Considered Doubtful",
             [("BS-AS-02-03-06"), ("BS-AS-02-03-07"), ("BS-AS-02-03-08"),
              ("BS-AS-02-03-09"), ("BS-AS-02-03-10")], False),
            ("Disputed Trade Receivables", None, True),
            ("  Considered Good",
             [("BS-AS-02-03-11"), ("BS-AS-02-03-12"), ("BS-AS-02-03-13"),
              ("BS-AS-02-03-14"), ("BS-AS-02-03-15")], False),
            ("  Considered Doubtful",
             [("BS-AS-02-03-16"), ("BS-AS-02-03-17"), ("BS-AS-02-03-18"),
              ("BS-AS-02-03-19"), ("BS-AS-02-03-20")], False),
            ("Others", [("BS-AS-02-03-21"), None, None, None, None], False),
            ("Total", None, True),
        ]

        data_start_row = r2 + 1
        for j, (label_row, codes, is_header) in enumerate(rows_data):
            row = r2 + 1 + j
            is_total = label_row == "Total"
            f = FONT_SECTION if (is_header or is_total) else FONT_ITEM
            fl = FILL_SECTION if is_header else (FILL_BLUE if is_total else FILL_WHITE)

            set_cell(ws, row, 1, label_row, font=f, fill=fl, border=BORDER_ALL, align=ALIGN_LEFT)

            if codes is None:  # section header or total
                for c in range(2, 8):
                    if is_total:
                        col = get_column_letter(c)
                        set_cell(ws, row, c,
                                 f"=SUM({col}{data_start_row}:{col}{row-1})",
                                 font=FONT_SECTION, fill=fl, border=BORDER_ALL, fmt=NF_AMOUNT)
                    else:
                        set_cell(ws, row, c, None, font=f, fill=fl, border=BORDER_ALL)
            else:
                # 5 bucket columns + total
                bucket_key = "cy_net" if label == "Current Reporting Period" else "py_net"
                cy_total = 0
                for c_idx, code in enumerate(codes[:5]):
                    col = c_idx + 2
                    if code:
                        val = balances.get(code, {}).get(bucket_key, 0)
                        cy_total += val
                        set_cell(ws, row, col, round(val, 2), font=FONT_PLAIN,
                                 border=BORDER_ALL, fmt=NF_AMOUNT)
                    else:
                        set_cell(ws, row, col, 0, font=FONT_PLAIN, border=BORDER_ALL, fmt=NF_AMOUNT)
                set_cell(ws, row, 7, round(cy_total, 2), font=FONT_PLAIN,
                         border=BORDER_ALL, fmt=NF_AMOUNT)

    set_column_widths(ws, {'A': 30, 'B': 15, 'C': 15, 'D': 15, 'E': 15, 'F': 15, 'G': 15})


def build_tp_ageing_sheet(wb, db: Session, project_id: int):
    """Sheet 12: Trade Payables Ageing (17 buckets)."""
    ws = wb.create_sheet("TP Ageing")
    setup_a4(ws, 'landscape')
    ws.sheet_properties.tabColor = "C62828"

    merge_title(ws, 1, _COMPANY_NAME, 6, FONT_TITLE)
    merge_title(ws, 2, "Trade Payables - Ageing Schedule", 6, FONT_SUBTITLE)

    from app.services.financial_engine import get_adjusted_balances
    balances = get_adjusted_balances(db, project_id)

    for label, start_row in [("Current Reporting Period", 5), ("Previous Reporting Period", 18)]:
        set_cell(ws, start_row, 1, f"Figures for the {label}", font=FONT_SECTION)
        r2 = start_row + 1
        header_row(ws, r2, [
            (1, "Particulars"), (2, "<1Y"), (3, "1-2Y"),
            (4, "2-3Y"), (5, ">3Y"), (6, "Total")
        ])

        bucket_key = "cy_net" if label == "Current Reporting Period" else "py_net"

        # MSME Undisputed: codes 01-04
        # Other Undisputed: 05-08
        # MSME Disputed: 09-12
        # Other Disputed: 13-16
        # Unbilled: 17
        rows_data = [
            ("(i) MSME Undisputed", ["BS-EL-04-02-01", "BS-EL-04-02-02",
                                      "BS-EL-04-02-03", "BS-EL-04-02-04"]),
            ("(ii) Others Undisputed", ["BS-EL-04-02-05", "BS-EL-04-02-06",
                                         "BS-EL-04-02-07", "BS-EL-04-02-08"]),
            ("(iii) MSME Disputed", ["BS-EL-04-02-09", "BS-EL-04-02-10",
                                      "BS-EL-04-02-11", "BS-EL-04-02-12"]),
            ("(iv) Others Disputed", ["BS-EL-04-02-13", "BS-EL-04-02-14",
                                       "BS-EL-04-02-15", "BS-EL-04-02-16"]),
            ("(v) Unbilled Dues", ["BS-EL-04-02-17", None, None, None]),
        ]

        data_start_row = r2 + 1
        for j, (label_row, codes) in enumerate(rows_data):
            row = r2 + 1 + j
            set_cell(ws, row, 1, label_row, font=FONT_ITEM, border=BORDER_ALL, align=ALIGN_LEFT)

            row_total = 0
            for c_idx, code in enumerate(codes):
                col = c_idx + 2
                if code:
                    val = -balances.get(code, {}).get(bucket_key, 0)  # Liability: negate
                    row_total += val
                    set_cell(ws, row, col, round(val, 2), font=FONT_PLAIN,
                             border=BORDER_ALL, fmt=NF_AMOUNT)
                else:
                    set_cell(ws, row, col, 0, font=FONT_PLAIN, border=BORDER_ALL, fmt=NF_AMOUNT)
            set_cell(ws, row, 6, round(row_total, 2), font=FONT_PLAIN,
                     border=BORDER_ALL, fmt=NF_AMOUNT)

        # Total row
        tr = r2 + 1 + len(rows_data)
        set_cell(ws, tr, 1, "Total", font=FONT_SECTION, fill=FILL_BLUE,
                 border=BORDER_ALL, align=ALIGN_LEFT)
        for c in range(2, 7):
            col = get_column_letter(c)
            set_cell(ws, tr, c, f"=SUM({col}{data_start_row}:{col}{tr-1})",
                     font=FONT_SECTION, fill=FILL_BLUE, border=BORDER_ALL, fmt=NF_AMOUNT)

    set_column_widths(ws, {'A': 32, 'B': 16, 'C': 16, 'D': 16, 'E': 16, 'F': 16})


def build_msme_sheet(wb):
    """Sheet 13: MSME Disclosure (skeleton)."""
    ws = wb.create_sheet("MSME")
    setup_a4(ws)
    ws.sheet_properties.tabColor = "F57F17"

    merge_title(ws, 1, _COMPANY_NAME, 4, FONT_TITLE)
    merge_title(ws, 2, "MSME Disclosure (MSMED Act 2006)", 4, FONT_SUBTITLE)

    r = 4
    header_row(ws, r, [(1, "S.No."), (2, "Particulars"),
                       (3, _BS_HEADER_CY), (4, _BS_HEADER_PY)])

    msme_items = [
        ("(a)", "Principal amount remaining unpaid at year end"),
        ("(a)", "Interest due thereon unpaid"),
        ("(b)", "Interest paid u/s 16 MSMED Act"),
        ("(b)", "Payment made beyond appointed day"),
        ("(c)", "Interest due and payable for delay period"),
        ("(d)", "Interest accrued and remaining unpaid at year end"),
        ("(e)", "Further interest due in succeeding years (u/s 23)"),
    ]
    for i, (sno, text) in enumerate(msme_items):
        r2 = r + 1 + i
        set_cell(ws, r2, 1, sno, font=FONT_ITEM, border=BORDER_ALL, align=ALIGN_CENTER)
        set_cell(ws, r2, 2, text, font=FONT_ITEM, border=BORDER_ALL, align=ALIGN_WRAP)
        set_cell(ws, r2, 3, None, font=FONT_INPUT, fill=FILL_INPUT,
                 border=BORDER_ALL, fmt=NF_AMOUNT)
        set_cell(ws, r2, 4, None, font=FONT_INPUT, fill=FILL_INPUT,
                 border=BORDER_ALL, fmt=NF_AMOUNT)

    set_column_widths(ws, {'A': 6, 'B': 65, 'C': 18, 'D': 18})


def build_ppe_sheet(wb, db=None, project_id=None):
    """Sheet 14: PPE & Intangible Reconciliation — with actual data if available."""
    ws = wb.create_sheet("PPE Schedule")
    setup_a4(ws, 'landscape')
    ws.sheet_properties.tabColor = "5D4037"

    merge_title(ws, 1, _COMPANY_NAME, 11, FONT_TITLE)
    merge_title(ws, 2, f"Property, Plant & Equipment — {_BS_HEADER_CY}", 11, FONT_SUBTITLE)

    r = 4
    header_row(ws, r, [
        (1, "Asset"), (2, "Gross Opening"), (3, "Additions"),
        (4, "Disposals"), (5, "Gross Closing"),
        (6, "Dep Opening"), (7, "Dep for Year"),
        (8, "Dep on Disp"), (9, "Dep Closing"),
        (10, "Net CY"), (11, "Net PY")
    ])

    # Try to get actual PPE data
    ppe_data = None
    if db and project_id:
        from app.services import ppe_service
        ppe_data = ppe_service.get_ppe_schedule(db, project_id)

    def write_asset_rows(rows, start_row):
        r2 = start_row
        for row in rows:
            set_cell(ws, r2, 1, row["asset_class"], font=FONT_ITEM, border=BORDER_ALL)
            for c, key in enumerate(["gross_opening","gross_additions","gross_disposals","gross_closing",
                                      "dep_opening","dep_for_year","dep_on_disposals","dep_closing",
                                      "net_cy","net_py"], 2):
                val = row.get(key, 0)
                is_auto = key in ("gross_closing", "dep_closing", "net_cy")
                set_cell(ws, r2, c, round(val, 2),
                         font=FONT_COMPUTED if is_auto else FONT_PLAIN,
                         border=BORDER_ALL, fmt=NF_AMOUNT)
            r2 += 1
        return r2

    def write_total_row(totals, row_num, label):
        set_cell(ws, row_num, 1, label, font=FONT_SECTION, fill=FILL_SECTION, border=BORDER_ALL)
        for c, key in enumerate(["gross_opening","gross_additions","gross_disposals","gross_closing",
                                  "dep_opening","dep_for_year","dep_on_disposals","dep_closing",
                                  "net_cy","net_py"], 2):
            set_cell(ws, row_num, c, round(totals.get(key, 0), 2),
                     font=FONT_SECTION, fill=FILL_SECTION, border=BORDER_ALL, fmt=NF_AMOUNT)

    if ppe_data:
        r2 = r + 1
        # Tangible header
        set_cell(ws, r2, 1, "TANGIBLE ASSETS", font=FONT_SECTION, fill=FILL_BLUE, border=BORDER_ALL)
        for c in range(2, 12):
            set_cell(ws, r2, c, None, fill=FILL_BLUE, border=BORDER_ALL)
        r2 += 1
        r2 = write_asset_rows(ppe_data["tangible"], r2)
        write_total_row(ppe_data["total_tangible"], r2, "TOTAL TANGIBLE")
        r2 += 2

        # Intangible header
        set_cell(ws, r2, 1, "INTANGIBLE ASSETS", font=FONT_SECTION, fill=FILL_BLUE, border=BORDER_ALL)
        for c in range(2, 12):
            set_cell(ws, r2, c, None, fill=FILL_BLUE, border=BORDER_ALL)
        r2 += 1
        r2 = write_asset_rows(ppe_data["intangible"], r2)
        write_total_row(ppe_data["total_intangible"], r2, "TOTAL INTANGIBLE")
        r2 += 2

        write_total_row(ppe_data["grand_total"], r2, "GRAND TOTAL")
    else:
        # Fallback: empty skeleton
        assets = ["Land","Buildings","Plant & Equipment","Furniture","Vehicles",
                  "Office Equipment","Computers","Other Tangible","TOTAL TANGIBLE",
                  "","Goodwill","Software","Trademarks","Other Intangible","TOTAL INTANGIBLE","","GRAND TOTAL"]
        for i, a in enumerate(assets):
            r2 = r + 1 + i
            if not a: continue
            is_total = "TOTAL" in a
            set_cell(ws, r2, 1, a, font=FONT_SECTION if is_total else FONT_ITEM,
                     fill=FILL_SECTION if is_total else FILL_WHITE, border=BORDER_ALL)
            for c in range(2, 12):
                set_cell(ws, r2, c, None, font=FONT_INPUT, fill=FILL_INPUT, border=BORDER_ALL, fmt=NF_AMOUNT)

    set_column_widths(ws, {'A': 22, 'B': 13, 'C': 13, 'D': 13, 'E': 13,
                           'F': 13, 'G': 13, 'H': 13, 'I': 13, 'J': 14, 'K': 14})


def build_related_party_sheet(wb):
    """Sheet 15: Related Party Transactions (skeleton)."""
    ws = wb.create_sheet("Related Party")
    setup_a4(ws, 'landscape')
    ws.sheet_properties.tabColor = "AD1457"

    merge_title(ws, 1, _COMPANY_NAME, 11, FONT_TITLE)
    merge_title(ws, 2, "Related Party Transactions (AS-18)", 11, FONT_SUBTITLE)

    r = 4
    section_row(ws, r, 1, "A. LIST OF RELATED PARTIES", 11)
    header_row(ws, r + 1, [(1, "Category"), (2, "Name"),
                           (3, "Relationship"), (4, "PAN/CIN")])
    for i in range(15):
        for c in range(1, 5):
            set_cell(ws, r + 2 + i, c, None, font=FONT_INPUT, fill=FILL_INPUT,
                     border=BORDER_ALL)

    r2 = r + 19
    section_row(ws, r2, 1, "B. TRANSACTION MATRIX", 11)
    header_row(ws, r2 + 1, [
        (1, "Transaction"), (2, "KMP CY"), (3, "KMP PY"),
        (4, "Relatives CY"), (5, "Relatives PY"),
        (6, "Entities CY"), (7, "Entities PY"),
        (8, "Hold/Sub CY"), (9, "Hold/Sub PY"),
        (10, "Total CY"), (11, "Total PY")
    ])
    txns = [
        "Purchase of Goods", "Sale of Goods", "Services Rendered", "Services Received",
        "Loans Given", "Loans Taken", "Interest Paid", "Interest Received",
        "Remuneration", "Sitting Fees", "Rent Paid", "Rent Received",
        "Commission", "Guarantee Given", "Outstanding Receivable", "Outstanding Payable"
    ]
    for i, t in enumerate(txns):
        row = r2 + 2 + i
        set_cell(ws, row, 1, t, font=FONT_ITEM, border=BORDER_ALL)
        for c in range(2, 10):
            set_cell(ws, row, c, None, font=FONT_INPUT, fill=FILL_INPUT,
                     border=BORDER_ALL, fmt=NF_AMOUNT)
        set_cell(ws, row, 10, f"=B{row}+D{row}+F{row}+H{row}",
                 font=FONT_COMPUTED, border=BORDER_ALL, fmt=NF_AMOUNT)
        set_cell(ws, row, 11, f"=C{row}+E{row}+G{row}+I{row}",
                 font=FONT_COMPUTED, border=BORDER_ALL, fmt=NF_AMOUNT)

    set_column_widths(ws, {'A': 24, 'B': 13, 'C': 13, 'D': 13, 'E': 13,
                           'F': 13, 'G': 13, 'H': 13, 'I': 13, 'J': 14, 'K': 14})


def build_policies_sheet(wb):
    """Sheet 16: Accounting Policies template."""
    ws = wb.create_sheet("Acc Policies")
    setup_a4(ws)
    ws.sheet_properties.tabColor = "37474F"

    merge_title(ws, 1, _COMPANY_NAME, 4, FONT_TITLE)
    merge_title(ws, 2, "Significant Accounting Policies (Note 1)", 4, FONT_SUBTITLE)

    policies = [
        ("1. Basis of Preparation",
         "The financial statements are prepared under historical cost convention on accrual basis, "
         "in accordance with Indian GAAP and Accounting Standards specified u/s 133 of "
         "Companies Act 2013 read with Rule 7 of Companies (Accounts) Rules 2014."),
        ("2. Use of Estimates", "[Customize per company]"),
        ("3. Revenue Recognition", "[Customize]"),
        ("4. Property, Plant and Equipment", "[Customize: Cost model, component accounting if applicable]"),
        ("5. Depreciation", "[Customize: WDV/SLM, useful life per Schedule II]"),
        ("6. Intangible Assets", "[Customize]"),
        ("7. Investments", "[Customize: Cost less diminution]"),
        ("8. Inventories", "[Customize: Lower of cost & NRV, FIFO/Weighted avg]"),
        ("9. Employee Benefits",
         "[Customize: Defined contribution - PF/ESI; Defined benefit - Gratuity; "
         "Short-term - Bonus/Leave]"),
        ("10. Borrowing Costs", "[Customize: Capitalize for qualifying assets per AS-16]"),
        ("11. Taxation", "[Customize: Current + Deferred tax approach]"),
        ("12. Provisions & Contingencies", "[Customize per AS-29]"),
        ("13. Foreign Currency", "[Customize per AS-11]"),
        ("14. Earnings Per Share",
         "Basic EPS = PAT / Weighted avg shares. Diluted EPS considers dilutive potential shares."),
        ("15. Cash Flow Statement", "Prepared using indirect method as per AS-3."),
        ("16. Segment Reporting", "[Customize if applicable per AS-17]"),
        ("17. Leases", "[Customize per AS-19]"),
        ("18. Impairment of Assets", "[Customize per AS-28]"),
    ]

    r = 4
    for title, body in policies:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        set_cell(ws, r, 1, title, font=FONT_SECTION, fill=FILL_SECTION, align=ALIGN_WRAP)
        r += 1
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        fill = FILL_YELLOW if "[Customize" in body else FILL_WHITE
        set_cell(ws, r, 1, body, font=FONT_ITEM, fill=fill, align=ALIGN_WRAP)
        ws.row_dimensions[r].height = 40
        r += 2

    set_column_widths(ws, {'A': 20, 'B': 25, 'C': 25, 'D': 25})


def build_addl_disclosures_sheet(wb):
    """Sheet 17: Additional Disclosures (A-T)."""
    ws = wb.create_sheet("Addl Disclosures")
    setup_a4(ws)
    ws.sheet_properties.tabColor = "1B5E20"

    merge_title(ws, 1, _COMPANY_NAME, 4, FONT_TITLE)
    merge_title(ws, 2, "Additional Disclosures - Schedule III", 4, FONT_SUBTITLE)

    r = 4
    header_row(ws, r, [(1, "Ref"), (2, "Particulars"),
                       (3, _BS_HEADER_CY), (4, _BS_HEADER_PY)])

    disclosures = [
        ("A", "CIF VALUE OF IMPORTS", True),
        ("(i)", "Raw Materials", False),
        ("(ii)", "Components & Spares", False),
        ("(iii)", "Capital Goods", False),
        ("B", "EXPENDITURE IN FOREIGN CURRENCY", True),
        ("(i)", "Royalty / Know-how", False),
        ("(ii)", "Professional Fees", False),
        ("(iii)", "Interest", False),
        ("(iv)", "Others", False),
        ("C", "EARNINGS IN FOREIGN EXCHANGE", True),
        ("(i)", "Export of Goods (FOB)", False),
        ("(ii)", "Professional Fees", False),
        ("(iii)", "Interest & Dividend", False),
        ("(iv)", "Others", False),
        ("D", "IMPORTED vs INDIGENOUS CONSUMPTION", True),
        ("", "Imported - Amount", False),
        ("", "Imported - %", False),
        ("", "Indigenous - Amount", False),
        ("", "Indigenous - %", False),
        ("E", "DIVIDENDS IN FOREIGN CURRENCY", True),
        ("", "No. of Non-Resident Shareholders", False),
        ("", "Shares Held", False),
        ("", "Amount Remitted", False),
        ("F", "SHARES HELD BY PROMOTERS (with % change)", True),
        ("", "(Table: Name, Shares, %, Change from PY)", False),
        ("G", "CWIP AGEING SCHEDULE", True),
        ("", "(<1Y, 1-2Y, 2-3Y, >3Y breakdown)", False),
        ("H", "INTANGIBLE UNDER DEVELOPMENT AGEING", True),
        ("", "(<1Y, 1-2Y, 2-3Y, >3Y breakdown)", False),
        ("I", "BENAMI PROPERTY", True),
        ("", "Details of proceedings / held property if any", False),
        ("J", "STRUCK-OFF COMPANIES (Sec 248)", True),
        ("", "Name / Nature / Balance", False),
        ("K", "CHARGES NOT REGISTERED WITH ROC", True),
        ("", "Beyond statutory period", False),
        ("L", "COMPLIANCE WITH LAYERS (Sec 2(87))", True),
        ("", "Compliance status", False),
        ("M", "UTILISATION OF BORROWED / SHARE PREMIUM FUNDS", True),
        ("", "Purpose other than for which taken", False),
        ("N", "UNDISCLOSED INCOME", True),
        ("", "Surrendered or disclosed during the year", False),
        ("O", "CRYPTO CURRENCY / VIRTUAL CURRENCY", True),
        ("", "Details of transactions", False),
        ("P", "WILFUL DEFAULTER", True),
        ("", "Declared as wilful defaulter by any Bank/FI", False),
        ("Q", "CSR DETAILED DISCLOSURE (Sec 135)", True),
        ("", "Amount required to be spent", False),
        ("", "Amount spent", False),
        ("", "Amount unspent - Ongoing projects", False),
        ("", "Amount unspent - Other than ongoing", False),
        ("R", "LOANS / ADVANCES (Sec 186)", True),
        ("", "Loans to promoters / directors / KMP / related parties", False),
        ("", "Investments by company in body corporate", False),
        ("", "Guarantee / security provided", False),
        ("S", "EVENTS AFTER BALANCE SHEET DATE", True),
        ("", "Material events requiring disclosure", False),
        ("T", "SEGMENT REPORTING", True),
        ("", "If applicable per AS-17", False),
    ]

    for i, (ref, text, is_section) in enumerate(disclosures):
        r2 = r + 1 + i
        set_cell(ws, r2, 1, ref,
                 font=FONT_SECTION if is_section else FONT_ITEM,
                 fill=FILL_SECTION if is_section else FILL_WHITE,
                 border=BORDER_ALL, align=ALIGN_CENTER)
        set_cell(ws, r2, 2, text,
                 font=FONT_SECTION if is_section else FONT_ITEM,
                 fill=FILL_SECTION if is_section else FILL_WHITE,
                 border=BORDER_ALL, align=ALIGN_WRAP)
        if not is_section:
            set_cell(ws, r2, 3, None, font=FONT_INPUT, fill=FILL_INPUT,
                     border=BORDER_ALL, fmt=NF_AMOUNT)
            set_cell(ws, r2, 4, None, font=FONT_INPUT, fill=FILL_INPUT,
                     border=BORDER_ALL, fmt=NF_AMOUNT)

    set_column_widths(ws, {'A': 5, 'B': 60, 'C': 18, 'D': 18})


def build_eps_sheet(wb, db: Session, project_id: int):
    """Sheet 18: EPS computation."""
    ws = wb.create_sheet("EPS")
    setup_a4(ws)
    ws.sheet_properties.tabColor = "4A148C"

    merge_title(ws, 1, _COMPANY_NAME, 5, FONT_TITLE)
    merge_title(ws, 2, "Earnings Per Share Computation (AS-20)", 5, FONT_SUBTITLE)

    eps = eps_engine.generate_eps(db, project_id)

    r = 4
    section_row(ws, r, 1, "PART A: PROFIT AVAILABLE TO EQUITY SHAREHOLDERS", 5)
    header_row(ws, r + 1, [(2, "Particulars"), (4, "CY"), (5, "PY")])

    eps_a = [
        ("Profit for the Period (PAT)", eps["pat_cy"], eps["pat_py"], True),
        ("Less: Preference Dividend (if any)", eps["preference_dividend"], eps["preference_dividend"], False),
        ("Profit Available to Equity Shareholders", eps["profit_to_equity_cy"], eps["profit_to_equity_py"], True),
    ]
    for i, (text, cy, py, is_total) in enumerate(eps_a):
        r2 = r + 2 + i
        f = FONT_SECTION if is_total else FONT_ITEM
        fl = FILL_SECTION if is_total else FILL_WHITE
        set_cell(ws, r2, 2, text, font=f, fill=fl, border=BORDER_ALL, align=ALIGN_LEFT)
        set_cell(ws, r2, 4, round(cy, 2), font=f, fill=fl, border=BORDER_ALL,
                 fmt=NF_AMOUNT, align=ALIGN_RIGHT)
        set_cell(ws, r2, 5, round(py, 2), font=f, fill=fl, border=BORDER_ALL,
                 fmt=NF_AMOUNT, align=ALIGN_RIGHT)

    r2 = r + 6
    section_row(ws, r2, 1, "PART B: WEIGHTED AVERAGE SHARES", 5)
    header_row(ws, r2 + 1, [(1, "Event"), (2, "Date"),
                            (3, "Shares"), (4, "Days Out"), (5, "Weighted")])
    for i in range(8):
        r3 = r2 + 2 + i
        for c in range(1, 6):
            set_cell(ws, r3, c, None, font=FONT_INPUT, fill=FILL_INPUT, border=BORDER_ALL)

    r3 = r2 + 12
    section_row(ws, r3, 1, "PART C: EPS CALCULATION", 5)
    header_row(ws, r3 + 1, [(2, "Particulars"), (4, "CY"), (5, "PY")])

    eps_c = [
        ("Weighted Avg Shares - Basic", eps["weighted_avg_basic"], eps["weighted_avg_basic"], False),
        ("Weighted Avg Shares - Diluted", eps["weighted_avg_diluted"], eps["weighted_avg_diluted"], False),
        ("", None, None, False),
        ("Basic EPS (Rs.)", eps["basic_eps_cy"], eps["basic_eps_py"], True),
        ("Diluted EPS (Rs.)", eps["diluted_eps_cy"], eps["diluted_eps_py"], True),
        ("Face Value per Share (Rs.)", eps["face_value"], eps["face_value"], False),
    ]
    for i, (text, cy, py, is_total) in enumerate(eps_c):
        r4 = r3 + 2 + i
        f = FONT_SECTION if is_total else FONT_ITEM
        fl = FILL_BLUE if is_total else FILL_WHITE
        if text:
            set_cell(ws, r4, 2, text, font=f, fill=fl, border=BORDER_ALL, align=ALIGN_LEFT)
            if cy is not None:
                set_cell(ws, r4, 4, cy, font=f, fill=fl, border=BORDER_ALL,
                         fmt=NF_AMOUNT_2, align=ALIGN_RIGHT)
                set_cell(ws, r4, 5, py, font=f, fill=fl, border=BORDER_ALL,
                         fmt=NF_AMOUNT_2, align=ALIGN_RIGHT)

    set_column_widths(ws, {'A': 18, 'B': 32, 'C': 18, 'D': 18, 'E': 18})


def build_ratios_sheet(wb, db: Session, project_id: int):
    """Sheet 19: Ratio Analysis (11 mandatory ratios)."""
    ws = wb.create_sheet("Ratios")
    setup_a4(ws, 'landscape')
    ws.sheet_properties.tabColor = "0D47A1"

    merge_title(ws, 1, _COMPANY_NAME, 12, FONT_TITLE)
    merge_title(ws, 2, "Ratio Analysis (Schedule III Requirement)", 12, FONT_SUBTITLE)

    r = 4
    header_row(ws, r, [
        (1, "#"), (2, "Ratio"), (3, "Numerator"),
        (4, "Num CY"), (5, "Num PY"),
        (6, "Denominator"), (7, "Den CY"), (8, "Den PY"),
        (9, "CY"), (10, "PY"), (11, "% Chg"), (12, "Explanation (if >25%)")
    ])

    rs = ratio_engine.generate_ratios(db, project_id)

    for i, ratio in enumerate(rs["ratios"]):
        r2 = r + 1 + i
        set_cell(ws, r2, 1, ratio["no"], font=FONT_ITEM, border=BORDER_ALL, align=ALIGN_CENTER)
        set_cell(ws, r2, 2, ratio["name"], font=FONT_SECTION, border=BORDER_ALL)
        set_cell(ws, r2, 3, ratio["numerator"], font=FONT_ITEM, border=BORDER_ALL)
        set_cell(ws, r2, 4, round(ratio["num_cy"], 2), font=FONT_PLAIN,
                 border=BORDER_ALL, fmt=NF_AMOUNT)
        set_cell(ws, r2, 5, round(ratio["num_py"], 2), font=FONT_PLAIN,
                 border=BORDER_ALL, fmt=NF_AMOUNT)
        set_cell(ws, r2, 6, ratio["denominator"], font=FONT_ITEM, border=BORDER_ALL)
        set_cell(ws, r2, 7, round(ratio["den_cy"], 2), font=FONT_PLAIN,
                 border=BORDER_ALL, fmt=NF_AMOUNT)
        set_cell(ws, r2, 8, round(ratio["den_py"], 2), font=FONT_PLAIN,
                 border=BORDER_ALL, fmt=NF_AMOUNT)
        set_cell(ws, r2, 9, ratio["cy"] if ratio["cy"] is not None else "-",
                 font=FONT_PLAIN, border=BORDER_ALL, fmt=NF_RATIO)
        set_cell(ws, r2, 10, ratio["py"] if ratio["py"] is not None else "-",
                 font=FONT_PLAIN, border=BORDER_ALL, fmt=NF_RATIO)
        set_cell(ws, r2, 11, ratio["variance_pct"] if ratio["variance_pct"] is not None else "-",
                 font=FONT_RED_BOLD if ratio["flag"] else FONT_PLAIN,
                 border=BORDER_ALL, fmt=NF_PCT)
        set_cell(ws, r2, 12, "FLAGGED - requires explanation" if ratio["flag"] else "",
                 font=FONT_RED_BOLD if ratio["flag"] else FONT_ITEM,
                 fill=FILL_YELLOW if ratio["flag"] else FILL_WHITE,
                 border=BORDER_ALL, align=ALIGN_WRAP)

    set_column_widths(ws, {
        'A': 5, 'B': 24, 'C': 20, 'D': 14, 'E': 14,
        'F': 20, 'G': 14, 'H': 14, 'I': 10, 'J': 10, 'K': 10, 'L': 30,
    })


def build_contingent_sheet(wb):
    """Sheet 20: Contingent Liabilities (skeleton)."""
    ws = wb.create_sheet("Contingent")
    setup_a4(ws)
    ws.sheet_properties.tabColor = "BF360C"

    merge_title(ws, 1, _COMPANY_NAME, 4, FONT_TITLE)
    merge_title(ws, 2, "Contingent Liabilities & Commitments", 4, FONT_SUBTITLE)

    r = 4
    header_row(ws, r, [(1, "Ref"), (2, "Particulars"),
                       (3, _BS_HEADER_CY), (4, _BS_HEADER_PY)])

    items = [
        ("I", "CONTINGENT LIABILITIES", True),
        ("(a)", "Claims against company not acknowledged as debt", False),
        ("(b)", "Guarantees", False),
        ("(c)", "Other contingent liabilities", False),
        ("", "Total Contingent Liabilities", True),
        ("II", "COMMITMENTS", True),
        ("(a)", "Capital contracts remaining to be executed", False),
        ("(b)", "Uncalled liability on investments", False),
        ("(c)", "Other commitments", False),
        ("", "Total Commitments", True),
    ]
    for i, (ref, text, is_section) in enumerate(items):
        r2 = r + 1 + i
        set_cell(ws, r2, 1, ref,
                 font=FONT_SECTION if is_section else FONT_ITEM,
                 fill=FILL_SECTION if is_section else FILL_WHITE,
                 border=BORDER_ALL, align=ALIGN_CENTER)
        set_cell(ws, r2, 2, text,
                 font=FONT_SECTION if is_section else FONT_ITEM,
                 fill=FILL_SECTION if is_section else FILL_WHITE,
                 border=BORDER_ALL)
        if not is_section:
            set_cell(ws, r2, 3, None, font=FONT_INPUT, fill=FILL_INPUT,
                     border=BORDER_ALL, fmt=NF_AMOUNT)
            set_cell(ws, r2, 4, None, font=FONT_INPUT, fill=FILL_INPUT,
                     border=BORDER_ALL, fmt=NF_AMOUNT)

    set_column_widths(ws, {'A': 5, 'B': 50, 'C': 18, 'D': 18})


# ============================================================
# MASTER BUILD FUNCTION
# ============================================================

def export_full_excel(db: Session, project_id: int, output_path: str = None) -> str:
    """
    Build the complete 20-sheet Excel for a project.
    Returns the filepath of the saved file.
    """
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise ValueError(f"Project {project_id} not found")
    client = project.client

    # Set module-level context for all sheet builders
    global _COMPANY_NAME, _BS_HEADER_CY, _BS_HEADER_PY, _PL_HEADER_CY, _PL_HEADER_PY, _SIGNING
    _COMPANY_NAME = client.name or "Company Name"
    _BS_HEADER_CY = project.bs_header_cy
    _BS_HEADER_PY = project.bs_header_py
    _PL_HEADER_CY = project.pl_header_cy
    _PL_HEADER_PY = project.pl_header_py

    # Load signing block
    from app.models import SigningBlock
    sb = db.query(SigningBlock).filter(SigningBlock.project_id == project_id).first()
    _SIGNING = {
        "auditor_firm": sb.auditor_firm, "auditor_frn": sb.auditor_frn,
        "partner_name": sb.partner_name, "partner_membership_no": sb.partner_membership_no,
        "partner_udin": sb.partner_udin,
        "director1_name": sb.director1_name, "director1_din": sb.director1_din,
        "director1_designation": sb.director1_designation,
        "director2_name": sb.director2_name, "director2_din": sb.director2_din,
        "director2_designation": sb.director2_designation,
        "place": sb.place, "signing_date": str(sb.signing_date) if sb.signing_date else "",
    } if sb else None

    wb = Workbook()

    # Build all sheets in order
    build_master_sheet(wb, client, project)           # 1
    build_tb_sheet(wb, db, project_id)                 # 2
    build_audit_sheet(wb, db, project_id)              # 3
    build_coa_sheet(wb, db)                             # 4
    build_balance_sheet(wb, db, project_id)            # 5
    build_pl_sheet(wb, db, project_id)                 # 6
    build_notes_bs_sheet(wb, db, project_id)           # 7
    build_notes_pl_sheet(wb, db, project_id)           # 8
    build_cashflow_sheet(wb, db, project_id)           # 9
    build_socie_sheet(wb)                               # 10
    build_tr_ageing_sheet(wb, db, project_id)          # 11
    build_tp_ageing_sheet(wb, db, project_id)          # 12
    build_msme_sheet(wb)                                # 13
    build_ppe_sheet(wb, db, project_id)                 # 14
    build_related_party_sheet(wb)                       # 15
    build_policies_sheet(wb)                            # 16
    build_addl_disclosures_sheet(wb)                    # 17
    build_eps_sheet(wb, db, project_id)                # 18
    build_ratios_sheet(wb, db, project_id)             # 19
    build_contingent_sheet(wb)                          # 20

    # Determine output path
    if not output_path:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_name = "".join(c if c.isalnum() or c in "._- " else "_" for c in client.name)[:40]
        filename = f"{safe_name}_{project.financial_year}_{timestamp}.xlsx"
        output_path = os.path.join(OUTPUT_DIR, filename)

    wb.save(output_path)
    return output_path
