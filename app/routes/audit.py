"""Audit routes - manage proposed audit adjustment entries"""
from datetime import date
from fastapi import APIRouter, Depends, HTTPException
from pydantic import BaseModel
from sqlalchemy.orm import Session
from app.database import get_db
from app.services import audit_service

router = APIRouter()


class AuditCreate(BaseModel):
    project_id: int
    dr_coa_code: str
    cr_coa_code: str
    amount: float
    narration: str = ""
    entry_date: date | None = None
    status: str = "proposed"


class AuditUpdate(BaseModel):
    narration: str | None = None
    dr_coa_code: str | None = None
    cr_coa_code: str | None = None
    amount: float | None = None
    status: str | None = None
    entry_date: date | None = None


@router.post("/audit/")
def create_audit(payload: AuditCreate, db: Session = Depends(get_db)):
    """Create new audit adjustment entry."""
    try:
        e = audit_service.create_audit_entry(
            db,
            project_id=payload.project_id,
            dr_coa_code=payload.dr_coa_code,
            cr_coa_code=payload.cr_coa_code,
            amount=payload.amount,
            narration=payload.narration,
            entry_date=payload.entry_date,
            status=payload.status,
        )
        return {
            "id": e.id,
            "entry_no": e.entry_no,
            "status": "created",
        }
    except ValueError as e:
        raise HTTPException(400, str(e))


@router.get("/audit/{project_id}")
def list_audit(project_id: int, status: str = None, db: Session = Depends(get_db)):
    """List audit entries for a project."""
    return audit_service.list_audit_entries(db, project_id, status_filter=status)


@router.put("/audit/{entry_id}")
def update_audit(entry_id: int, payload: AuditUpdate, db: Session = Depends(get_db)):
    """Update an audit entry."""
    try:
        updates = payload.model_dump(exclude_none=True)
        # Map entry_date to 'date' field
        if 'entry_date' in updates:
            updates['date'] = updates.pop('entry_date')
        e = audit_service.update_audit_entry(db, entry_id, **updates)
        return {"id": e.id, "status": "updated"}
    except ValueError as e:
        raise HTTPException(404, str(e))


@router.get("/audit/{project_id}/export")
def export_audit_entries(project_id: int, db: Session = Depends(get_db)):
    """Export audit entries as Excel."""
    import openpyxl, io
    from fastapi.responses import StreamingResponse
    from app.models import AuditEntry, CoAMaster

    entries = db.query(AuditEntry).filter(AuditEntry.project_id == project_id).order_by(AuditEntry.entry_no).all()
    coa = {c.code: c.particulars for c in db.query(CoAMaster).all()}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Audit Entries"
    headers = ["#", "Date", "Narration", "Dr Code", "Dr Particulars", "Cr Code", "Cr Particulars", "Amount", "Status"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill("solid", fgColor="DAEEF3")

    for e in entries:
        ws.append([
            e.entry_no or '', e.date.strftime("%d-%b-%Y") if e.date else '',
            e.narration, e.dr_coa_code, coa.get(e.dr_coa_code, ''),
            e.cr_coa_code, coa.get(e.cr_coa_code, ''), e.amount, e.status
        ])

    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=Audit_Entries_{project_id}.xlsx"})


@router.delete("/audit/{entry_id}")
def delete_audit(entry_id: int, db: Session = Depends(get_db)):
    """Delete an audit entry."""
    try:
        return audit_service.delete_audit_entry(db, entry_id)
    except ValueError as e:
        raise HTTPException(404, str(e))


@router.get("/audit/{project_id}/check")
def audit_check(project_id: int, db: Session = Depends(get_db)):
    """Audit totals balance check."""
    return audit_service.audit_totals_check(db, project_id)
