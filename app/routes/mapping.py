"""Mapping routes - auto-map TB ledgers to CoA codes"""
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session
from app.database import get_db
from app.services import mapping_service
from app.models import CoAMaster, TrialBalance
import io

router = APIRouter()


class ManualMap(BaseModel):
    tb_row_id: int
    coa_code: str


class BatchMap(BaseModel):
    project_id: int
    mappings: list[ManualMap]


@router.post("/map/{project_id}/auto")
def auto_map(project_id: int, force: bool = False, db: Session = Depends(get_db)):
    """Auto-map all TB ledgers for a project."""
    try:
        return mapping_service.auto_map_project(db, project_id, force=force)
    except Exception as e:
        raise HTTPException(500, str(e))


@router.post("/map/manual")
def manual_map(payload: ManualMap, db: Session = Depends(get_db)):
    """Manually assign a CoA code to a TB row."""
    try:
        return mapping_service.set_manual_mapping(db, payload.tb_row_id, payload.coa_code)
    except ValueError as e:
        raise HTTPException(404, str(e))


@router.post("/map/batch")
def batch_map(payload: BatchMap, db: Session = Depends(get_db)):
    """C4: Save all mappings in one go."""
    saved = 0
    for m in payload.mappings:
        try:
            mapping_service.set_manual_mapping(db, m.tb_row_id, m.coa_code)
            saved += 1
        except Exception:
            pass
    return {"saved": saved, "total": len(payload.mappings)}


@router.get("/export-mapped-tb/{project_id}")
def export_mapped_tb(project_id: int, db: Session = Depends(get_db)):
    """C8: Download mapped TB as Excel with CoA codes."""
    import openpyxl
    rows = db.query(TrialBalance).filter(TrialBalance.project_id == project_id).order_by(TrialBalance.id).all()
    if not rows:
        raise HTTPException(404, "No TB data")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mapped TB"
    headers = ["Ledger", "Tally Group", "CoA Code", "CY Debit", "CY Credit", "CY Net", "PY Debit", "PY Credit", "PY Net"]
    ws.append(headers)
    for h in ws[1]:
        h.font = openpyxl.styles.Font(bold=True)

    for r in rows:
        ws.append([r.ledger_name, r.tally_group, r.coa_code or "",
                    r.cy_debit, r.cy_credit, r.cy_net,
                    r.py_debit, r.py_credit, r.py_net])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename=Mapped_TB_{project_id}.xlsx"})


@router.post("/import-mapped-tb/{project_id}")
async def import_mapped_tb(project_id: int, file: UploadFile = File(...), db: Session = Depends(get_db)):
    """C9: Import mapped TB — update CoA codes from uploaded file."""
    import openpyxl
    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents))
    ws = wb.active

    # Find header row
    headers = [str(c.value or "").strip().lower() for c in ws[1]]
    ledger_col = next((i for i, h in enumerate(headers) if h in ("ledger", "ledger_name", "ledger name")), None)
    code_col = next((i for i, h in enumerate(headers) if h in ("coa code", "coa_code", "code", "mapping")), None)
    if ledger_col is None or code_col is None:
        raise HTTPException(400, "Need columns: 'Ledger' and 'CoA Code'")

    # Load valid codes
    valid_codes = {c.code for c in db.query(CoAMaster.code).all()}
    rows = db.query(TrialBalance).filter(TrialBalance.project_id == project_id).all()
    ledger_map = {r.ledger_name.strip().lower(): r for r in rows}

    updated = 0
    skipped = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        ledger = str(row[ledger_col] or "").strip()
        code = str(row[code_col] or "").strip()
        if not ledger or not code:
            continue
        tb_row = ledger_map.get(ledger.lower())
        if tb_row and code in valid_codes:
            tb_row.coa_code = code
            updated += 1
        else:
            skipped += 1

    db.commit()
    return {"updated": updated, "skipped": skipped}


@router.get("/map/{project_id}/summary")
def mapping_summary(project_id: int, db: Session = Depends(get_db)):
    """Get mapping completion statistics."""
    return mapping_service.get_mapping_summary(db, project_id)


@router.get("/coa/")
def list_coa(fs_type: str = None, db: Session = Depends(get_db)):
    """List all CoA master codes (optionally filter by BS/PL)."""
    q = db.query(CoAMaster)
    if fs_type:
        q = q.filter(CoAMaster.fs_type == fs_type.upper())
    codes = q.order_by(CoAMaster.code).all()
    return [
        {
            "code": c.code,
            "level": c.level,
            "particulars": c.particulars,
            "schedule_ref": c.schedule_ref,
            "nature": c.nature,
            "fs_type": c.fs_type,
            "note_ref": c.note_ref,
            "tally_group": c.tally_group,
        }
        for c in codes
    ]


# ============= C5: Copy mapping between ledgers =============

class CopyMapping(BaseModel):
    source_tb_row_id: int
    target_tb_row_ids: list[int]


@router.post("/map/copy")
def copy_mapping(payload: CopyMapping, db: Session = Depends(get_db)):
    """C5: Copy CoA code from one ledger to multiple others."""
    source = db.query(TrialBalance).filter(TrialBalance.id == payload.source_tb_row_id).first()
    if not source or not source.coa_code:
        raise HTTPException(400, "Source row has no mapping")

    copied = 0
    for tid in payload.target_tb_row_ids:
        row = db.query(TrialBalance).filter(TrialBalance.id == tid).first()
        if row:
            row.coa_code = source.coa_code
            copied += 1
    db.commit()
    return {"copied": copied, "coa_code": source.coa_code}


# ============= C6: Import mapping from previous FY project =============

class ImportPrevMapping(BaseModel):
    source_project_id: int
    target_project_id: int


@router.post("/map/import-previous")
def import_previous_mapping(payload: ImportPrevMapping, db: Session = Depends(get_db)):
    """C6: Copy mappings from a previous FY project by matching ledger names."""
    from app.models import Project
    src_proj = db.query(Project).filter(Project.id == payload.source_project_id).first()
    tgt_proj = db.query(Project).filter(Project.id == payload.target_project_id).first()
    if not src_proj or not tgt_proj:
        raise HTTPException(404, "Project not found")

    # Build ledger→code map from source
    src_rows = db.query(TrialBalance).filter(TrialBalance.project_id == payload.source_project_id).all()
    src_map = {r.ledger_name.strip().lower(): r.coa_code for r in src_rows if r.coa_code}

    if not src_map:
        raise HTTPException(400, f"Source project (FY {src_proj.financial_year}) has no mappings")

    # Apply to target
    tgt_rows = db.query(TrialBalance).filter(TrialBalance.project_id == payload.target_project_id).all()
    matched = 0
    skipped = 0
    for r in tgt_rows:
        key = r.ledger_name.strip().lower()
        if key in src_map:
            r.coa_code = src_map[key]
            matched += 1
        else:
            skipped += 1
    db.commit()
    return {
        "matched": matched,
        "skipped": skipped,
        "source_fy": src_proj.financial_year,
        "source_total_mappings": len(src_map),
    }


@router.get("/map/previous-projects/{project_id}")
def list_previous_projects(project_id: int, db: Session = Depends(get_db)):
    """C6: List other projects of the same client for mapping import."""
    from app.models import Project
    proj = db.query(Project).filter(Project.id == project_id).first()
    if not proj:
        raise HTTPException(404, "Project not found")
    others = db.query(Project).filter(
        Project.client_id == proj.client_id,
        Project.id != project_id,
    ).all()
    result = []
    for p in others:
        mapped = db.query(TrialBalance).filter(
            TrialBalance.project_id == p.id, TrialBalance.coa_code.isnot(None)
        ).count()
        result.append({"id": p.id, "financial_year": p.financial_year,
                        "version": p.version or 1, "mapped_count": mapped})
    return result
