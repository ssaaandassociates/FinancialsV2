"""Upload route - TB file upload"""
from fastapi import APIRouter, UploadFile, File, Form, Depends, HTTPException
from sqlalchemy.orm import Session
from app.database import get_db
from app.services import tb_service

router = APIRouter()


@router.post("/upload-tb/{project_id}")
async def upload_tb(
    project_id: int,
    file: UploadFile = File(...),
    replace: bool = Form(True),
    db: Session = Depends(get_db),
):
    """Upload Trial Balance file (Excel or CSV) for a project."""
    try:
        file_bytes = await file.read()
        tb_df = tb_service.parse_tb_file(file_bytes, file.filename)
        result = tb_service.save_tb_to_db(db, project_id, tb_df, replace=replace)
        return {
            "project_id": project_id,
            "filename": file.filename,
            "parsed_rows": len(tb_df),
            **result,
        }
    except ValueError as e:
        raise HTTPException(400, str(e))
    except Exception as e:
        raise HTTPException(500, f"Error processing file: {str(e)}")


@router.get("/tb/{project_id}")
def get_tb(project_id: int, unmapped_only: bool = False, db: Session = Depends(get_db)):
    """Get TB data for a project."""
    if unmapped_only:
        return tb_service.get_unmapped_ledgers(db, project_id)
    return tb_service.get_tb_for_project(db, project_id)


@router.get("/tb-template/tally")
def download_tally_template():
    """Download sample Tally TB template (Excel)."""
    import openpyxl
    from fastapi.responses import FileResponse
    import os

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Trial Balance"
    headers = ["Ledger", "Tally Group", "CY Debit", "CY Credit", "PY Debit", "PY Credit"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    # Sample rows
    samples = [
        ("Share Capital", "Capital Account", 0, 100000, 0, 100000),
        ("Sales", "Sales Accounts", 0, 500000, 0, 400000),
        ("Bank Account", "Bank Accounts", 200000, 0, 150000, 0),
    ]
    for r, row in enumerate(samples, 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)

    path = os.path.join("output", "TB_Template_Tally.xlsx")
    os.makedirs("output", exist_ok=True)
    wb.save(path)
    return FileResponse(path, filename="TB_Template_Tally.xlsx",
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@router.get("/tb-template/generic")
def download_generic_template():
    """Download generic TB template (no Tally group, user provides CoA code)."""
    import openpyxl
    from fastapi.responses import FileResponse
    import os

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Trial Balance"
    headers = ["Ledger", "CoA Code", "CY Debit", "CY Credit", "PY Debit", "PY Credit"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    samples = [
        ("Share Capital", "BS-EL-01-01-02", 0, 100000, 0, 100000),
        ("Sales - Products", "PL-01-01-01", 0, 500000, 0, 400000),
        ("HDFC Bank", "BS-AS-02-04-01", 200000, 0, 150000, 0),
    ]
    for r, row in enumerate(samples, 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)

    # Add CoA reference sheet — populated with all standard leaf codes so the
    # user can copy the exact code into the "CoA Code" column.
    ws2 = wb.create_sheet("CoA Reference")
    ws2.cell(1, 1, "Code")
    ws2.cell(1, 2, "Particulars")
    ws2.cell(1, 3, "Nature")
    ws2.cell(1, 4, "FS Type")
    try:
        from app.database import SessionLocal
        from app.models import CoAMaster
        _db = SessionLocal()
        try:
            codes = _db.query(CoAMaster).order_by(CoAMaster.code).all()
            for r, cc in enumerate(codes, 2):
                ws2.cell(r, 1, cc.code)
                ws2.cell(r, 2, cc.particulars)
                ws2.cell(r, 3, getattr(cc, "nature", "") or "")
                ws2.cell(r, 4, getattr(cc, "fs_type", "") or "")
        finally:
            _db.close()
    except Exception:
        ws2.cell(2, 1, "(Could not load CoA list — see app for codes)")
    # widen columns
    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 45

    path = os.path.join("output", "TB_Template_Generic.xlsx")
    os.makedirs("output", exist_ok=True)
    wb.save(path)
    return FileResponse(path, filename="TB_Template_Generic.xlsx",
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
