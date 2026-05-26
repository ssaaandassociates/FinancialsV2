"""Supplementary data routes - Ageing, Share Capital, Related Party, Disclosures, Policies"""
from datetime import date
from fastapi import APIRouter, UploadFile, File, Form, Depends, HTTPException
from pydantic import BaseModel
from sqlalchemy.orm import Session
from app.database import get_db
from app.services import ageing_engine, share_capital_engine, related_party_engine, disclosure_engine

router = APIRouter()


# ============= AGEING =============

class AgeingEntry(BaseModel):
    project_id: int
    ageing_type: str  # TR or TP
    party_name: str
    is_msme: bool = False
    is_disputed: bool = False
    is_doubtful: bool = False
    bucket_1: float = 0
    bucket_2: float = 0
    bucket_3: float = 0
    bucket_4: float = 0
    bucket_5: float = 0
    period: str = "CY"


@router.post("/ageing/")
def add_ageing(payload: AgeingEntry, db: Session = Depends(get_db)):
    """Add a single party ageing entry."""
    try:
        buckets = {f"bucket_{i}": getattr(payload, f"bucket_{i}") for i in range(1, 6)}
        entry = ageing_engine.save_ageing_entry(
            db, payload.project_id, payload.ageing_type, payload.party_name,
            buckets, payload.period, payload.is_msme, payload.is_disputed, payload.is_doubtful
        )
        return {"id": entry.id, "status": "created"}
    except Exception as e:
        raise HTTPException(400, str(e))


@router.post("/ageing/{project_id}/upload")
async def upload_ageing(
    project_id: int,
    ageing_type: str = Form(...),
    period: str = Form("CY"),
    replace: bool = Form(True),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """Upload ageing data from Excel/CSV."""
    try:
        file_bytes = await file.read()
        result = ageing_engine.upload_ageing_file(
            db, project_id, file_bytes, file.filename,
            ageing_type, period, replace
        )
        return result
    except ValueError as e:
        raise HTTPException(400, str(e))


@router.get("/ageing/{project_id}")
def get_ageing(project_id: int, ageing_type: str = "TR",
               period: str = "CY", db: Session = Depends(get_db)):
    """Get ageing data for a project."""
    return ageing_engine.get_ageing_data(db, project_id, ageing_type, period)


@router.get("/ageing/{project_id}/schedule/tr")
def tr_schedule(project_id: int, db: Session = Depends(get_db)):
    """Generate TR ageing schedule."""
    return ageing_engine.generate_tr_ageing_schedule(db, project_id)


@router.get("/ageing/{project_id}/schedule/tp")
def tp_schedule(project_id: int, db: Session = Depends(get_db)):
    """Generate TP ageing schedule."""
    return ageing_engine.generate_tp_ageing_schedule(db, project_id)


# ============= SHARE CAPITAL =============

class ShareEventInput(BaseModel):
    project_id: int
    event_type: str  # opening / issue / bonus / rights / buyback / forfeiture
    event_date: date | None = None
    share_class: str = "equity"
    no_of_shares: int = 0
    face_value: float = 10
    premium: float = 0
    period: str = "CY"
    narration: str = ""


class ShareholderInput(BaseModel):
    project_id: int
    name: str
    no_of_shares_cy: int = 0
    no_of_shares_py: int = 0
    share_class: str = "equity"
    is_promoter: bool = False
    din_pan: str = ""


@router.post("/share-events/")
def add_share_event(payload: ShareEventInput, db: Session = Depends(get_db)):
    """Add a share capital movement event."""
    try:
        event = share_capital_engine.save_share_event(
            db, payload.project_id, payload.event_type,
            payload.event_date, payload.share_class,
            payload.no_of_shares, payload.face_value, payload.premium,
            payload.period, payload.narration
        )
        return {"id": event.id, "type": event.event_type, "shares": event.no_of_shares}
    except Exception as e:
        raise HTTPException(400, str(e))


@router.post("/shareholders/")
def add_shareholder(payload: ShareholderInput, db: Session = Depends(get_db)):
    """Add a shareholder record."""
    try:
        sh = share_capital_engine.save_shareholder(
            db, payload.project_id, payload.name,
            payload.no_of_shares_cy, payload.no_of_shares_py,
            payload.share_class, payload.is_promoter, payload.din_pan
        )
        return {"id": sh.id, "name": sh.name}
    except Exception as e:
        raise HTTPException(400, str(e))


@router.get("/share-capital/{project_id}/reconciliation")
def share_reconciliation(project_id: int, db: Session = Depends(get_db)):
    """Get share capital reconciliation."""
    return share_capital_engine.get_share_reconciliation(db, project_id)


@router.get("/share-capital/{project_id}/shareholders")
def shareholders(project_id: int, db: Session = Depends(get_db)):
    """Get all shareholders with >5% flagged."""
    return share_capital_engine.get_major_shareholders(db, project_id)


@router.get("/share-capital/{project_id}/promoters")
def promoter_holding(project_id: int, db: Session = Depends(get_db)):
    """Get promoter shareholding with % change."""
    return share_capital_engine.get_promoter_holding(db, project_id)


@router.get("/share-capital/{project_id}/full")
def full_note_a(project_id: int, db: Session = Depends(get_db)):
    """Generate complete Note A: Share Capital disclosure."""
    return share_capital_engine.generate_full_note_a(db, project_id)


# ============= RELATED PARTY =============

class RPInput(BaseModel):
    project_id: int
    name: str
    category: str  # KMP / Relative of KMP / Entity / Holding/Sub
    relationship: str = ""
    pan_cin: str = ""


class RPTxnInput(BaseModel):
    project_id: int
    party_id: int
    transaction_type: str
    cy_amount: float = 0
    py_amount: float = 0


@router.post("/related-parties/")
def add_rp(payload: RPInput, db: Session = Depends(get_db)):
    """Add a related party."""
    try:
        p = related_party_engine.add_related_party(
            db, payload.project_id, payload.name,
            payload.category, payload.relationship, payload.pan_cin
        )
        return {"id": p.id, "name": p.name, "category": p.category}
    except Exception as e:
        raise HTTPException(400, str(e))


@router.get("/related-parties/{project_id}")
def list_rp(project_id: int, db: Session = Depends(get_db)):
    """List all related parties for a project."""
    return related_party_engine.list_related_parties(db, project_id)


@router.post("/rp-transactions/")
def add_rp_txn(payload: RPTxnInput, db: Session = Depends(get_db)):
    """Add/update a related party transaction."""
    try:
        txn = related_party_engine.add_rp_transaction(
            db, payload.project_id, payload.party_id,
            payload.transaction_type, payload.cy_amount, payload.py_amount
        )
        return {"id": txn.id, "type": txn.transaction_type, "cy": txn.cy_amount}
    except Exception as e:
        raise HTTPException(400, str(e))


@router.get("/rp-transactions/party/{party_id}")
def list_party_transactions(party_id: int, db: Session = Depends(get_db)):
    """List all transactions for one related party (Section 7 V9)."""
    return related_party_engine.list_transactions_for_party(db, party_id)


class RPTxnUpdate(BaseModel):
    transaction_type: str | None = None
    cy_amount: float | None = None
    py_amount: float | None = None


@router.put("/rp-transactions/{txn_id}")
def update_rp_txn(txn_id: int, payload: RPTxnUpdate, db: Session = Depends(get_db)):
    """Update a single transaction (Section 7 V9)."""
    txn = related_party_engine.update_rp_transaction(
        db, txn_id,
        transaction_type=payload.transaction_type,
        cy_amount=payload.cy_amount,
        py_amount=payload.py_amount,
    )
    if not txn:
        raise HTTPException(404, "Transaction not found")
    return {"id": txn.id, "type": txn.transaction_type,
            "cy": txn.cy_amount, "py": txn.py_amount}


@router.delete("/rp-transactions/{txn_id}")
def delete_rp_txn(txn_id: int, db: Session = Depends(get_db)):
    """Delete a single transaction (Section 7 V9)."""
    ok = related_party_engine.delete_rp_transaction(db, txn_id)
    if not ok:
        raise HTTPException(404, "Transaction not found")
    return {"status": "deleted"}


@router.get("/rp/kmp-candidates/{project_id}")
def kmp_candidates(project_id: int, db: Session = Depends(get_db)):
    """
    Section 7 V9: Return candidate related parties from client master
    (all directors + non-director shareholders). For autocomplete dropdown.
    """
    from app.models import Project
    proj = db.query(Project).filter(Project.id == project_id).first()
    if not proj:
        raise HTTPException(404, "Project not found")
    return related_party_engine.get_kmp_candidates_for_client(db, proj.client_id)


@router.get("/related-parties/{project_id}/disclosure")
def rp_disclosure(project_id: int, db: Session = Depends(get_db)):
    """Generate full RP disclosure with transaction matrix."""
    return related_party_engine.generate_rp_disclosure(db, project_id)


@router.delete("/related-parties/{party_id}")
def delete_rp(party_id: int, db: Session = Depends(get_db)):
    """Delete a related party."""
    from app.models.supplementary import RelatedParty, RPTransaction
    db.query(RPTransaction).filter(RPTransaction.party_id == party_id).delete()
    db.query(RelatedParty).filter(RelatedParty.id == party_id).delete()
    db.commit()
    return {"status": "deleted"}


@router.post("/rp/auto-kmp/{project_id}")
def auto_add_kmp_as_rp(project_id: int, db: Session = Depends(get_db)):
    """D4: Auto-add directors with is_kmp=True as Related Parties."""
    from app.models import Project
    from app.models.client import Director
    from app.models.supplementary import RelatedParty

    proj = db.query(Project).filter(Project.id == project_id).first()
    if not proj:
        raise HTTPException(404, "Project not found")

    directors = db.query(Director).filter(
        Director.client_id == proj.client_id,
        Director.is_active == True,
        Director.is_kmp == True,
    ).all()

    existing = {rp.name.strip().lower() for rp in
                db.query(RelatedParty).filter(RelatedParty.project_id == project_id).all()}

    added = 0
    for d in directors:
        if d.name.strip().lower() not in existing:
            rp = RelatedParty(
                project_id=project_id,
                name=d.name,
                category="KMP",
                relationship=d.designation or "Director",
                pan_cin=d.pan or "",
            )
            db.add(rp)
            added += 1
    db.commit()
    return {"added": added, "total_kmp": len(directors)}


# ============= D6: AGEING TEMPLATES =============

@router.get("/ageing-template/{ageing_type}")
def download_ageing_template(ageing_type: str):
    """D6: Download blank ageing template (TR or TP)."""
    import openpyxl
    from fastapi.responses import StreamingResponse
    import io

    wb = openpyxl.Workbook()
    ws = wb.active

    if ageing_type.upper() == "TR":
        ws.title = "TR Ageing"
        headers = ["Party Name", "Disputed (Y/N)", "Doubtful (Y/N)",
                    "<6 Months", "6M-1Y", "1-2Y", "2-3Y", ">3Y"]
        ws.append(headers)
        # Sample row
        ws.append(["ABC Traders", "N", "N", 50000, 20000, 10000, 5000, 2000])
    else:
        ws.title = "TP Ageing"
        headers = ["Party Name", "MSME (Y/N)", "Disputed (Y/N)",
                    "<1 Year", "1-2Y", "2-3Y", ">3Y"]
        ws.append(headers)
        ws.append(["XYZ Suppliers", "N", "N", 80000, 15000, 5000, 1000])

    # Format header
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill("solid", fgColor="DAEEF3")

    # Auto column widths
    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = max(max_len + 2, 12)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={ageing_type.upper()}_Ageing_Template.xlsx"},
    )


# ============= ACCOUNTING POLICIES =============

class PolicyUpdate(BaseModel):
    title: str | None = None
    body: str | None = None
    is_active: bool | None = None


class PolicyCreate(BaseModel):
    project_id: int
    title: str
    body: str


@router.get("/policies/{project_id}")
def get_policies(project_id: int, db: Session = Depends(get_db)):
    """Get accounting policies (auto-seeds defaults on first call)."""
    return disclosure_engine.get_policies(db, project_id)


@router.put("/policies/{policy_id}")
def update_policy(policy_id: int, payload: PolicyUpdate, db: Session = Depends(get_db)):
    """Update a policy."""
    try:
        p = disclosure_engine.update_policy(db, policy_id, **payload.model_dump(exclude_none=True))
        return {"id": p.id, "title": p.title, "status": "updated"}
    except ValueError as e:
        raise HTTPException(404, str(e))


@router.post("/policies/")
def add_policy(payload: PolicyCreate, db: Session = Depends(get_db)):
    """Add a custom policy."""
    p = disclosure_engine.add_custom_policy(db, payload.project_id, payload.title, payload.body)
    return {"id": p.id, "number": p.policy_number, "title": p.title}


# ============= ADDITIONAL DISCLOSURES =============

class DisclosureUpdate(BaseModel):
    cy_amount: float | None = None
    py_amount: float | None = None
    text_value: str | None = None
    particulars: str | None = None


@router.get("/disclosures/{project_id}")
def get_disclosures(project_id: int, db: Session = Depends(get_db)):
    """Get all additional disclosures (auto-seeds structure on first call)."""
    return disclosure_engine.get_disclosures(db, project_id)


@router.put("/disclosures/{item_id}")
def update_disclosure(item_id: int, payload: DisclosureUpdate, db: Session = Depends(get_db)):
    """Update a disclosure item."""
    try:
        d = disclosure_engine.update_disclosure(db, item_id, **payload.model_dump(exclude_none=True))
        return {"id": d.id, "ref": d.disclosure_ref, "status": "updated"}
    except ValueError as e:
        raise HTTPException(404, str(e))
