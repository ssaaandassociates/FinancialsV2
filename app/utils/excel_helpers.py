"""
Excel formatting helpers - centralized styles, fonts, borders, page setup.
Used by export_engine.py for all sheet builders.
"""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.page import PageMargins

# ---------- COLOR PALETTE ----------
NAVY = "1A3A6B"
GREEN = "1A9E6B"
ORANGE = "E65100"
BLUE = "1565C0"
RED = "F44336"
GREY = "999999"

# ---------- BORDERS ----------
thin = Side(style='thin', color=GREY)
hair = Side(style='hair', color='CCCCCC')
medium = Side(style='medium', color=NAVY)

BORDER_ALL = Border(left=thin, right=thin, top=thin, bottom=thin)
BORDER_HAIR = Border(left=hair, right=hair, top=hair, bottom=hair)
BORDER_BOTTOM_MEDIUM = Border(bottom=medium)

# ---------- FONTS ----------
FONT_TITLE = Font(name='Arial', bold=True, size=11, color=NAVY)
FONT_SUBTITLE = Font(name='Arial', bold=True, size=10, color=GREEN)
FONT_HEADER = Font(name='Arial', bold=True, size=9, color='FFFFFF')
FONT_SECTION = Font(name='Arial', bold=True, size=9, color=NAVY)
FONT_ITEM = Font(name='Arial', size=9, color='333333')
FONT_ITEM_LIGHT = Font(name='Arial', size=9, color='555555')
FONT_INPUT = Font(name='Arial', size=9, color='0000FF')
FONT_COMPUTED = Font(name='Arial', size=9, color='008000')
FONT_PLAIN = Font(name='Arial', size=9, color='000000')
FONT_NOTE = Font(name='Arial', italic=True, size=8, color='888888')
FONT_SMALL = Font(name='Arial', size=8, color='666666')
FONT_RED_BOLD = Font(name='Arial', bold=True, size=9, color='FF0000')

# ---------- FILLS ----------
FILL_HEADER = PatternFill('solid', fgColor=NAVY)
FILL_SECTION = PatternFill('solid', fgColor='E8EDF5')
FILL_INPUT = PatternFill('solid', fgColor='EBF5FB')
FILL_YELLOW = PatternFill('solid', fgColor='FFFFF0')
FILL_WHITE = PatternFill('solid', fgColor='FFFFFF')
FILL_LIGHT = PatternFill('solid', fgColor='F5F5F5')
FILL_GREEN = PatternFill('solid', fgColor='E8F8F0')
FILL_BLUE = PatternFill('solid', fgColor='F0F4FA')

# ---------- ALIGNMENTS ----------
ALIGN_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal='right', vertical='center')
ALIGN_WRAP = Alignment(vertical='top', wrap_text=True)

# ---------- NUMBER FORMATS ----------
NF_AMOUNT = '#,##0;(#,##0);"-"'
NF_AMOUNT_2 = '#,##0.00;(#,##0.00);"-"'
NF_PCT = '0.00%'
NF_RATIO = '0.00'


def setup_a4(ws, orientation='portrait'):
    """Set up A4 page with proper margins."""
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = orientation
    ws.page_margins = PageMargins(
        left=0.5, right=0.5, top=0.75, bottom=0.5,
        header=0.3, footer=0.3
    )
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.oddHeader.center.text = '&"Arial,Bold"&9=Master!B5'


def set_cell(ws, row, col, value, font=None, fill=None, align=None, fmt=None, border=None):
    """Set cell with all properties in one call."""
    cell = ws.cell(row=row, column=col, value=value)
    if font: cell.font = font
    if fill: cell.fill = fill
    if align: cell.alignment = align
    if fmt: cell.number_format = fmt
    if border: cell.border = border
    return cell


def merge_title(ws, row, text, merge_cols, font=None):
    """Merge row 1 to merge_cols and set title text."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=merge_cols)
    set_cell(ws, row, 1, text,
             font=font or FONT_TITLE,
             align=ALIGN_CENTER)


def header_row(ws, row, cols_with_labels):
    """cols_with_labels = [(col_num, label), ...]"""
    for col, label in cols_with_labels:
        set_cell(ws, row, col, label,
                 font=FONT_HEADER, fill=FILL_HEADER,
                 align=ALIGN_CENTER, border=BORDER_ALL)


def section_row(ws, row, col, text, merge_cols):
    """Section divider row - fills background across columns."""
    for c in range(1, merge_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = FILL_SECTION
        cell.border = BORDER_ALL
    set_cell(ws, row, col, text,
             font=FONT_SECTION, fill=FILL_SECTION,
             align=ALIGN_LEFT, border=BORDER_ALL)


def set_column_widths(ws, widths):
    """widths = {'A': 28, 'B': 18, ...}"""
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width
